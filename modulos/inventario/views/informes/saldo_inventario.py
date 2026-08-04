import io
import os
from typing import Any
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.articulos import Articulos
from modulos.maestros.models.bodegas import Bodegas
from modulos.maestros.models.auxiliares import TipoArticulo


class IndexInformeSaldoInvetarioView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/inventario/informes/saldo_inventario.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        if action == "buscar_articulo":
            return self._buscar_articulo(request.POST.get("codigo"))
        elif action == "listar_articulos":
            return self._listar_articulos()
        elif action == "generar_pdf":
            return self._generar_pdf(request)
        elif action == "generar_excel":
            return self._generar_excel(request)
        elif action == "listar_tipos":
            return self._listar_tipos()
        elif action == "generar_pdf_global":
            return self._generar_pdf_global(request)
        elif action == "generar_excel_global":
            return self._generar_excel_global(request)
        return JsonResponse({"success": False})

    def _buscar_articulo(self, codigo: str | None) -> JsonResponse:
        if not codigo:
            return JsonResponse({"success": False})
        try:
            art = Articulos.objects.get(codigo=codigo.strip())
            return JsonResponse({
                "success": True,
                "data": {
                    "codigo": art.codigo,
                    "nombre": art.descr or "",
                    "um": art.um or "",
                }
            })
        except Articulos.DoesNotExist:
            return JsonResponse({"success": False, "message": "Artículo no encontrado"})

    def _listar_articulos(self) -> JsonResponse:
        articulos = Articulos.objects.values("codigo", "descr", "um").order_by("descr")
        return JsonResponse({"articulos": list(articulos)})

    def _listar_tipos(self) -> JsonResponse:
        tipos = TipoArticulo.objects.values("id", "nombre").order_by("nombre")
        return JsonResponse({"tipos": list(tipos)})

    def _get_datos_globales(self, fecha_corte: str, tipo_art_id: str | None, solo_con_stock: bool):
        fc = None
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte.strip(), "%Y-%m-%d")
            except Exception:
                pass

        base_qs = (
            Movs.objects
            .select_related("tipo", "codigo", "codigo__tipo_articulo")
            .filter(
                codigo__isnull=False,
                tipo__isnull=False,
            )
            .exclude(tipo__cod__in=[8, 15])
            .exclude(codigo__tipo='Servicio')
            .exclude(codigo__codigo='')
        )

        if fc:
            base_qs = base_qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        if tipo_art_id:
            base_qs = base_qs.filter(codigo__tipo_articulo__id=tipo_art_id)

        base_qs = base_qs.distinct().order_by("codigo", "fecha")

        data: dict[str, dict] = {}
        for m in base_qs.iterator():
            if not m.codigo:
                continue
            cod = m.codigo.codigo
            if cod not in data:
                ta = m.codigo.tipo_articulo
                tipo_nombre = ta.nombre if ta else (m.codigo.tipo.strip() or "Sin tipo")
                data[cod] = {
                    "nombre": m.codigo.descr or "",
                    "um": m.codigo.um or "",
                    "saldo": 0,
                    "ultimo_precio": m.codigo.cup or 0,
                    "tipo_nombre": tipo_nombre,
                }
            data[cod]["saldo"] += m.cantidad or 0

        resultados = []
        for codigo, info in data.items():
            if solo_con_stock and info["saldo"] == 0:
                continue
            resultados.append({
                "codigo": codigo,
                "nombre": info["nombre"],
                "um": info["um"],
                "saldo": info["saldo"],
                "ultimo_precio": info["ultimo_precio"],
                "costo": info["ultimo_precio"] * info["saldo"],
                "tipo_nombre": info["tipo_nombre"],
            })

        resultados.sort(key=lambda x: (x["tipo_nombre"], x["codigo"]))
        return resultados

    def _get_movimientos(self, codigo: str, fecha_corte: str):
        fc = None
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte.strip(), "%Y-%m-%d")
            except Exception:
                pass

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(
                codigo__codigo=codigo,
                tipo__isnull=False,
            )
            .exclude(tipo__cod__in=[8, 15])
            .exclude(codigo__tipo='Servicio')
        )

        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.distinct().order_by("fecha")
        return qs, fc

    def _generar_excel(self, request: HttpRequest) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        codigo = request.POST.get("codigo", "").strip()
        fecha_corte = request.POST.get("fecha_corte", "").strip()

        if not codigo:
            return JsonResponse({"success": False, "message": "Debe ingresar un código de artículo"})

        qs, fc = self._get_movimientos(codigo, fecha_corte)

        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}

        rows = []
        total_entradas = 0
        total_salidas = 0
        ultimo_precio = 0
        for m in qs.iterator():
            signo = m.tipo.signo if m.tipo else 0
            cantidad = m.cantidad or 0
            if signo > 0:
                total_entradas += cantidad
            if signo < 0:
                total_salidas += abs(cantidad)
            if m.punit:
                ultimo_precio = m.punit
            c_stock = (m.cantidad or 0) * (m.punit or 0)
            rows.append([
                m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                cantidad,
                bodegas_map.get(m.bodega, str(m.bodega or "")),
                str(int(m.numero)) if m.numero else "",
                m.tipo.nombre if m.tipo else "",
                m.punit or 0,
                c_stock,
            ])

        saldo_final = total_entradas - total_salidas
        total_costo = ultimo_precio * saldo_final

        wb = Workbook()
        ws = wb.active
        ws.title = "Saldo Inventario"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Fecha", "Cantidad", "Bodega", "Número", "Documento", "P.Unitario", "C.Stock"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        for r in rows:
            ws.append(r)
            row_num = ws.max_row
            for col_idx in range(1, len(r) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in (2, 6, 7):
                    cell.alignment = num_align
                    if col_idx == 2:
                        cell.number_format = '#,##0.000'
                    else:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = center_align

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 32
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16

        ws.append([])

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        ws.append([f"Total Entradas: {clq(total_entradas)}  |  Total Salidas: {clq(total_salidas)}  |  Saldo Final: {clq(saldo_final)}  |  Total Costo: {cl(total_costo)}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"saldo_inventario_{codigo}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf(self, request: HttpRequest) -> HttpResponse:
        codigo = request.POST.get("codigo", "").strip()
        fecha_corte = request.POST.get("fecha_corte", "").strip()
        usuario = str(request.user)

        if not codigo:
            return JsonResponse({"success": False, "message": "Debe ingresar un código de artículo"})

        qs, fc = self._get_movimientos(codigo, fecha_corte)

        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}

        rows = []
        total_entradas = 0
        total_salidas = 0
        ultimo_precio = 0

        for m in qs.iterator():
            signo = m.tipo.signo if m.tipo else 0
            cantidad = m.cantidad or 0
            if signo > 0:
                total_entradas += cantidad
            if signo < 0:
                total_salidas += abs(cantidad)
            if m.punit:
                ultimo_precio = m.punit
            c_stock = (m.cantidad or 0) * (m.punit or 0)
            rows.append([
                m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                cantidad,
                bodegas_map.get(m.bodega, str(m.bodega or "")),
                str(int(m.numero)) if m.numero else "",
                m.tipo.nombre if m.tipo else "",
                m.punit or 0,
                c_stock,
            ])

        saldo_final = total_entradas - total_salidas
        total_costo = ultimo_precio * saldo_final

        try:
            art_info = Articulos.objects.get(codigo=codigo)
            articulo_nombre = art_info.descr or ""
            um = art_info.um or ""
        except Articulos.DoesNotExist:
            articulo_nombre = ""
            um = ""

        fecha_inf = ""
        if fc:
            fecha_inf = f"al {fc.strftime('%d-%m-%Y')}"

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=7, leading=9,
            )
            right_style = ParagraphStyle(
                "RightStyle", parent=cell_style, alignment=2,
            )
            center_style = ParagraphStyle(
                "CenterStyle", parent=cell_style, alignment=1,
            )
            resumen_style = ParagraphStyle(
                "Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1,
            )

            elems = []
            elems.append(Paragraph("Saldo de Inventario por Artículo", title_style))
            elems.append(Paragraph(
                f"<b>Código:</b> {codigo} &nbsp;&nbsp;&nbsp;"
                f"<b>Nombre:</b> {articulo_nombre} &nbsp;&nbsp;&nbsp;"
                f"<b>UM:</b> {um}",
                subtitle_style,
            ))
            if fecha_inf:
                elems.append(Paragraph(f"<b>Corte:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 3 * mm))

            header = ["Fecha", "Cantidad", "Bodega", "Número", "Documento", "P.Unitario", "C.Stock"]
            table_data = [header]

            for r in rows:
                table_data.append([
                    Paragraph(str(r[0]), center_style),
                    Paragraph(f"<b>{clq(r[1])}</b>", right_style),
                    Paragraph(str(r[2]), cell_style),
                    Paragraph(str(r[3]), center_style),
                    Paragraph(str(r[4]), center_style),
                    Paragraph(cl(r[5]), right_style),
                    Paragraph(cl(r[6]), right_style),
                ])

            col_widths = [26 * mm, 20 * mm, 42 * mm, 22 * mm, 40 * mm, 20 * mm, 22 * mm]

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]))

            elems.append(tbl)
            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[
                Paragraph(f'<b>Total Entradas</b><br/><font color="#16a34a">{clq(total_entradas)}</font>', resumen_style),
                Paragraph(f'<b>Total Salidas</b><br/><font color="#dc2626">{clq(total_salidas)}</font>', resumen_style),
                Paragraph(f'<b>Saldo Final</b><br/><font color="#2563eb">{clq(saldo_final)}</font>', resumen_style),
                Paragraph(f'<b>Total Costo</b><br/><font color="#7c3aed">{cl(total_costo)}</font>', resumen_style),
            ]]

            resumen_tbl = Table(resumen_data, colWidths=[32 * mm, 32 * mm, 32 * mm, 32 * mm], hAlign="CENTER")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))

            elems.append(resumen_tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=22 * mm, bottomMargin=18 * mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = f"saldo_inventario_{codigo}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    def _generar_excel_global(self, request: HttpRequest) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        fecha_corte = request.POST.get("fecha_corte", "").strip()
        tipo_art_raw = request.POST.get("tipo", "").strip()
        tipo_art_id = int(tipo_art_raw) if tipo_art_raw.isdigit() else None
        solo_con_stock = request.POST.get("solo_con_stock") == "true"
        print(f"[SALDO GLOBAL EXCEL] tipo_art_raw={tipo_art_raw!r}, tipo_art_id={tipo_art_id}, solo_con_stock={solo_con_stock}")

        data = self._get_datos_globales(fecha_corte, tipo_art_id, solo_con_stock)

        wb = Workbook()
        ws = wb.active
        ws.title = "Saldo Inventario Global"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Código", "Nombre", "UM", "Saldo", "Último Precio", "Costo"]
        num_cols = len(headers)
        ws.append(headers)
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        section_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        section_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        section_align = Alignment(horizontal="left", vertical="center")

        data_font = Font(name="Calibri", size=10)
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        total_saldo = 0
        total_costo = 0
        tipo_actual = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        for r in data:
            if r["tipo_nombre"] != tipo_actual:
                tipo_actual = r["tipo_nombre"]
                ws.append([f"▸ {tipo_actual}"] + [""] * (num_cols - 1))
                section_row = ws.max_row
                for col_idx in range(1, num_cols + 1):
                    cell = ws.cell(row=section_row, column=col_idx)
                    cell.font = section_font
                    cell.fill = section_fill
                    cell.alignment = section_align

            ws.append([r["codigo"], r["nombre"], r["um"], r["saldo"], r["ultimo_precio"], r["costo"]])
            total_saldo += r["saldo"]
            total_costo += r["costo"]
            row_num = ws.max_row
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if col_idx == 4:
                    cell.alignment = num_align
                    cell.number_format = '#,##0.000'
                elif col_idx in (5, 6):
                    cell.alignment = num_align
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = center_align

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16

        ws.append([])
        ws.append([f"Total Artículos: {len(data)}  |  Total Saldo: {clq(total_saldo)}  |  Total Costo: {cl(total_costo)}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "saldo_inventario_global.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf_global(self, request: HttpRequest) -> HttpResponse:
        fecha_corte = request.POST.get("fecha_corte", "").strip()
        tipo_art_raw = request.POST.get("tipo", "").strip()
        tipo_art_id = int(tipo_art_raw) if tipo_art_raw.isdigit() else None
        solo_con_stock = request.POST.get("solo_con_stock") == "true"
        usuario = str(request.user)
        print(f"[SALDO GLOBAL] tipo_art_raw={tipo_art_raw!r}, tipo_art_id={tipo_art_id}, solo_con_stock={solo_con_stock}, fecha_corte={fecha_corte!r}")

        data = self._get_datos_globales(fecha_corte, tipo_art_id, solo_con_stock)

        fc = None
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte.strip(), "%Y-%m-%d")
            except Exception:
                pass

        fecha_inf = ""
        if fc:
            fecha_inf = f"al {fc.strftime('%d-%m-%Y')}"

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=7, leading=9,
            )
            right_style = ParagraphStyle(
                "RightStyle", parent=cell_style, alignment=2,
            )
            center_style = ParagraphStyle(
                "CenterStyle", parent=cell_style, alignment=1,
            )
            resumen_style = ParagraphStyle(
                "Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1,
            )

            elems = []
            elems.append(Paragraph("Saldo de Inventario Global por Artículo", title_style))
            if fecha_inf:
                elems.append(Paragraph(f"<b>Corte:</b> {fecha_inf}", subtitle_style))

            if tipo_art_id:
                tipo_nombre = ""
                try:
                    tipo_nombre = TipoArticulo.objects.get(id=tipo_art_id).nombre
                except (TipoArticulo.DoesNotExist, ValueError):
                    pass
                if tipo_nombre:
                    elems.append(Paragraph(f"<b>Tipo Artículo:</b> {tipo_nombre}", subtitle_style))

            if solo_con_stock:
                elems.append(Paragraph("<b>Filtro:</b> Solo artículos con stock", subtitle_style))

            elems.append(Spacer(1, 3 * mm))

            header = ["Código", "Nombre", "UM", "Saldo", "P.Unitario", "Costo"]
            col_widths = [24 * mm, 72 * mm, 14 * mm, 24 * mm, 24 * mm, 24 * mm]

            tipo_actual = None
            sub_total_saldo = 0
            sub_total_costo = 0
            sub_count = 0
            table_data = None
            total_saldo = 0
            total_costo = 0

            section_bg = colors.HexColor("#374151")
            section_fg = colors.white
            section_style = ParagraphStyle(
                "SectionTitle", parent=styles["Normal"], fontSize=8, leading=10,
                textColor=section_fg, fontName="Helvetica-Bold",
            )
            center_section = ParagraphStyle(
                "CenterSection", parent=section_style, alignment=1,
            )

            def flush_subtotal():
                nonlocal sub_total_saldo, sub_total_costo, sub_count
                if sub_count > 0:
                    sub_style = ParagraphStyle(
                        "Subtotal", parent=styles["Normal"], fontSize=7, leading=9,
                        textColor=colors.HexColor("#4b5563"), fontName="Helvetica-Bold",
                    )
                    rs = ParagraphStyle("Rsub", parent=sub_style, alignment=2)
                    table_data.append([
                        Paragraph("<b>Subtotal</b>", sub_style),
                        Paragraph("", sub_style),
                        Paragraph("", sub_style),
                        Paragraph(f"<b>{clq(sub_total_saldo)}</b>", rs),
                        Paragraph("", sub_style),
                        Paragraph(f"<b>{cl(sub_total_costo)}</b>", rs),
                    ])
                    sub_total_saldo = 0
                    sub_total_costo = 0
                    sub_count = 0

            def start_section(name):
                nonlocal table_data
                table_data = []

            def end_section():
                if table_data and len(table_data) > 0:
                    nonlocal sub_total_saldo, sub_total_costo, sub_count
                    flush_subtotal()
                    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
                    tbl.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ]))
                    elems.append(tbl)
                    elems.append(Spacer(1, 3 * mm))

            for r in data:
                if r["tipo_nombre"] != tipo_actual:
                    end_section()
                    tipo_actual = r["tipo_nombre"]
                    table_data = [[
                        Paragraph(f"▸ {tipo_actual}", center_section),
                        Paragraph("", section_style), Paragraph("", section_style),
                        Paragraph("", section_style), Paragraph("", section_style),
                        Paragraph("", section_style),
                    ]]
                    sub_total_saldo = 0
                    sub_total_costo = 0
                    sub_count = 0

                sub_count += 1
                if len(table_data) == 1:
                    table_data.append(header)
                table_data.append([
                    Paragraph(str(r["codigo"]), center_style),
                    Paragraph(str(r["nombre"]), cell_style),
                    Paragraph(str(r["um"]), center_style),
                    Paragraph(f"<b>{clq(r['saldo'])}</b>", right_style),
                    Paragraph(cl(r["ultimo_precio"]), right_style),
                    Paragraph(cl(r["costo"]), right_style),
                ])
                sub_total_saldo += r["saldo"]
                sub_total_costo += r["costo"]
                total_saldo += r["saldo"]
                total_costo += r["costo"]
            end_section()

            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[
                Paragraph(f'<b>Total Artículos</b><br/><font color="#1f2937">{len(data)}</font>', resumen_style),
                Paragraph(f'<b>Total Saldo</b><br/><font color="#2563eb">{clq(total_saldo)}</font>', resumen_style),
                Paragraph(f'<b>Total Costo</b><br/><font color="#7c3aed">{cl(total_costo)}</font>', resumen_style),
            ]]

            resumen_tbl = Table(resumen_data, colWidths=[54 * mm, 54 * mm, 54 * mm], hAlign="CENTER")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))

            elems.append(resumen_tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=22 * mm, bottomMargin=18 * mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = "saldo_inventario_global.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response
