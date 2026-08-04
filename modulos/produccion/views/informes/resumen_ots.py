import io
import os
from datetime import datetime
from typing import Any

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.empleados import Empleados
from modulos.maestros.models.procesos import Procesos
from modulos.maestros.models.articulos import Articulos


class IndexInformeResumenOtsView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/produccion/informes/resumen_ots.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        if action == "listar_ots":
            return self._listar_ots(request)
        elif action == "listar_encargados":
            return self._listar_encargados()
        elif action == "cargar_subformularios":
            return self._cargar_subformularios(request)
        elif action == "buscar_ot_por_numero":
            return self._buscar_ot_por_numero(request)
        elif action == "generar_pdf":
            return self._generar_pdf(request)
        elif action == "generar_pdf_detalle":
            return self._generar_pdf_detalle(request)
        elif action == "generar_excel":
            return self._generar_excel(request)
        return JsonResponse({"success": False})

    def _listar_encargados(self) -> JsonResponse:
        empleados = Empleados.objects.values("cod", "nombre").order_by("nombre")
        return JsonResponse({"encargados": list(empleados)})

    def _get_fecha_params(self, request):
        fecha_inicio = request.POST.get("fecha_inicio", "").strip()
        fecha_fin = request.POST.get("fecha_fin", "").strip()
        
        fi = None
        ff = None
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d")
            except Exception:
                pass
        if fecha_fin:
            try:
                ff = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d")
            except Exception:
                pass
        
        return fi, ff

    def _listar_ots(self, request: HttpRequest) -> JsonResponse:
        try:
            qs = Movs.objects.filter(linea=0, tipo=8).values(
                "numero", "fecha", "codencargado", "proceso", "estado"
            ).order_by("-numero")

            procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}
            empleados_map = {e["cod"]: e["nombre"] for e in Empleados.objects.values("cod", "nombre")}

            ot_list = []
            for m in qs:
                numero = int(m["numero"]) if m["numero"] else 0
                if numero == 0:
                    continue

                proc_val = int(m["proceso"]) if m["proceso"] else 0
                enc_val = int(m["codencargado"]) if m["codencargado"] else 0

                ot_list.append({
                    "numero": numero,
                    "fecha": m["fecha"].strftime("%d-%m-%Y") if m["fecha"] else "",
                    "proceso": procesos_map.get(proc_val, ""),
                    "encargado": empleados_map.get(enc_val, ""),
                    "estado": m["estado"] or "",
                })

            return JsonResponse({"ots": ot_list})
        except Exception as e:
            return JsonResponse({"ots": [], "error": str(e)})

    def _cargar_subformularios(self, request: HttpRequest) -> JsonResponse:
        ot_num = request.POST.get("ot", "").strip()

        if not ot_num:
            return JsonResponse({"detalle_ot": [], "vale_consumo": [], "parte_entrada": [], "error": "sin ot"})

        try:
            ot_val = float(ot_num)
        except ValueError:
            return JsonResponse({"detalle_ot": [], "vale_consumo": [], "parte_entrada": [], "error": "invalid ot"})

        def get_detalle_ot():
            qs = Movs.objects.select_related("codigo").filter(
                tipo__cod__in=[8],
                numero=ot_val
            ).exclude(linea=0)
            rows = []
            for m in qs.iterator():
                cantidad = abs(m.cantidad) if m.cantidad else 0
                if cantidad == 0:
                    continue
                rows.append({
                    "codigo": m.codigo.codigo if m.codigo else "",
                    "nombre": m.codigo.descr if m.codigo else "",
                    "cantidad": cantidad,
                    "um": m.codigo.um if m.codigo else "",
                    "punit": m.punit or 0,
                    "bodega": str(m.bodega) if m.bodega else "",
                    "fecha": m.fecha.strftime("%Y-%m-%d") if m.fecha else "",
                    "estado": m.estado or "",
                    "docref": m.docref or "",
                    "rut": m.rut or "",
                    "canttotal": m.canttotal or 0,
                    "tipodocref": m.tipodocref or "",
                    "codencargado": m.codencargado or "",
                })
            return rows

        def get_vale_consumo():
            qs = Movs.objects.select_related("codigo").filter(
                tipo__cod__in=[10],
                numero=ot_val,
                linea__gt=0
            )
            rows = []
            for m in qs.iterator():
                cantidad = abs(m.cantidad) if m.cantidad else 0
                if cantidad == 0:
                    continue
                rows.append({
                    "codigo": m.codigo.codigo if m.codigo else "",
                    "nombre": m.codigo.descr if m.codigo else "",
                    "cantidad": cantidad,
                    "um": m.codigo.um if m.codigo else "",
                    "fecha": m.fecha.strftime("%Y-%m-%d") if m.fecha else "",
                    "codencargado": m.codencargado or "",
                })
            return rows

        def get_parte_entrada():
            qs = Movs.objects.select_related("codigo").filter(
                tipo__cod__in=[6],
                numero=ot_val,
                linea__gt=0
            )
            rows = []
            for m in qs.iterator():
                cantidad = abs(m.cantidad) if m.cantidad else 0
                if cantidad == 0:
                    continue
                rows.append({
                    "codigo": m.codigo.codigo if m.codigo else "",
                    "nombre": m.codigo.descr if m.codigo else "",
                    "cantidad": cantidad,
                    "um": m.codigo.um if m.codigo else "",
                    "fecha": m.fecha.strftime("%Y-%m-%d") if m.fecha else "",
                    "codencargado": m.codencargado or "",
                })
            return rows

        return JsonResponse({
            "detalle_ot": get_detalle_ot(),
            "vale_consumo": get_vale_consumo(),
            "parte_entrada": get_parte_entrada(),
        })

    def _buscar_ot_por_numero(self, request: HttpRequest) -> JsonResponse:
        ot_num = request.POST.get("ot", "").strip()

        if not ot_num:
            return JsonResponse({"found": False})

        try:
            ot_val = float(ot_num)
        except ValueError:
            return JsonResponse({"found": False})

        qs = Movs.objects.filter(tipo=8, numero=ot_val, linea=0).first()

        if not qs:
            return JsonResponse({"found": False})

        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}
        empleados_map = {e["cod"]: e["nombre"] for e in Empleados.objects.values("cod", "nombre")}

        proc_nom = procesos_map.get(int(qs.proceso), "") if qs.proceso else ""
        enc_nom = empleados_map.get(int(qs.codencargado), "") if qs.codencargado else ""

        return JsonResponse({
            "found": True,
            "numero": int(qs.numero) if qs.numero else 0,
            "fecha": qs.fecha.strftime("%d-%m-%Y") if qs.fecha else "",
            "encargado": enc_nom,
            "proceso": proc_nom,
            "estado": qs.estado or "",
            "glosa": qs.glosa or "",
            "rut": qs.rut or "",
        })

    def _get_datos(self, request: HttpRequest):
        fi, ff = self._get_fecha_params(request)
        ot_num = request.POST.get("ot", "").strip()

        qs = Movs.objects.select_related("tipo").filter(tipo__cod__in=[8]).exclude(linea=0).exclude(numero__isnull=True).exclude(numero=0)

        if fi:
            qs = qs.filter(fecha__gte=fi)
        if ff:
            qs = qs.filter(fecha__lte=ff.replace(hour=23, minute=59, second=59))
        if ot_num:
            try:
                qs = qs.filter(numero=float(ot_num))
            except ValueError:
                pass

        qs = qs.order_by("fecha", "numero")

        return qs

    def _generar_excel(self, request: HttpRequest) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        qs = self._get_datos(request)
        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}
        empleados_map = {e["cod"]: e["nombre"] for e in Empleados.objects.values("cod", "nombre")}

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen OTs"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["N° OT", "Proceso", "Encargado", "Fecha", "Cantidad"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        current_fecha = None
        current_ot = None
        ot_subtotal = 0
        fecha_subtotal = 0

        def emit_subtotal_ot():
            nonlocal ot_subtotal
            if ot_subtotal > 0:
                ws.append(["", "Subtotal OT", "", "", ot_subtotal])
                row_num = ws.max_row
                for col_idx in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = thin_border
                    cell.font = Font(bold=True)
                ot_subtotal = 0

        def emit_subtotal_fecha():
            nonlocal fecha_subtotal
            if fecha_subtotal > 0:
                ws.append(["", "", "", "Subtotal Fecha", fecha_subtotal])
                row_num = ws.max_row
                for col_idx in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = thin_border
                    cell.font = Font(bold=True)
                cell = ws.cell(row=row_num, column=4)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                fecha_subtotal = 0

        for m in qs.iterator():
            numero = int(m.numero) if m.numero else 0
            if numero == 0:
                continue

            fecha_key = m.fecha.strftime("%Y-%m-%d") if m.fecha else ""
            proceso_val = int(m.proceso) if m.proceso else 0
            enc_val = int(m.codencargado) if m.codencargado else 0
            cantidad = m.cantidad if m.cantidad else 0

            if current_fecha is not None and fecha_key != current_fecha:
                emit_subtotal_ot()
                emit_subtotal_fecha()

            if current_ot is not None and numero != current_ot:
                emit_subtotal_ot()

            current_fecha = fecha_key
            current_ot = numero

            ws.append([
                numero,
                procesos_map.get(proc_val, ""),
                empleados_map.get(enc_val, ""),
                m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                cantidad,
            ])
            ot_subtotal += cantidad
            fecha_subtotal += cantidad

        emit_subtotal_ot()
        emit_subtotal_fecha()

        row_num = ws.max_row
        for r in range(2, row_num + 1):
            for col_idx in range(1, 6):
                cell = ws.cell(row=r, column=col_idx)
                cell.border = thin_border
                if col_idx == 5 and ws.cell(row=r, column=1).value:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.000'
                elif col_idx == 1 and ws.cell(row=r, column=1).value:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 15

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "resumen_ots.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf(self, request: HttpRequest) -> HttpResponse:
        return self._generar_pdf_base(request, False)

    def _generar_pdf_res_pe(self, request: HttpRequest) -> HttpResponse:
        return self._generar_pdf_base(request, True)

    def _generar_pdf_base(self, request: HttpRequest, es_res_pe: bool) -> HttpResponse:
        fi, ff = self._get_fecha_params(request)
        ot_num = request.POST.get("ot", "").strip()
        usuario = str(request.user)

        fecha_inf = ""
        if fi:
            fecha_inf += f"del {fi.strftime('%d-%m-%Y')}"
        if ff:
            fecha_inf += f" al {ff.strftime('%d-%m-%Y')}"

        if es_res_pe:
            qs = Movs.objects.select_related("codigo", "tipo").filter(tipo__cod__in=[6], linea=1)
            titulo = "Producción por OT"
            header = ["Artículo", "Nombre", "Cantidad", "UM"]
            col_widths = [30 * mm, 80 * mm, 40 * mm, 30 * mm]
        else:
            qs = Movs.objects.select_related("tipo").filter(tipo__cod__in=[8]).exclude(linea=0)
            titulo = "Resumen de Órdenes de Trabajo"
            header = ["N° OT", "Proceso", "Encargado", "Fecha", "Cantidad"]
            col_widths = [20 * mm, 45 * mm, 35 * mm, 25 * mm, 30 * mm]

        if fi:
            qs = qs.filter(fecha__gte=fi)
        if ff:
            qs = qs.filter(fecha__lte=ff.replace(hour=23, minute=59, second=59))
        if ot_num:
            try:
                qs = qs.filter(numero=float(ot_num))
            except ValueError:
                pass

        qs = qs.order_by("fecha", "numero")
        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}
        empleados_map = {e["cod"]: e["nombre"] for e in Empleados.objects.values("cod", "nombre")}

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/brand-logos/logo-home-grande.png")
        if not os.path.exists(logo_path):
            logo_path = None

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")
        
        def clt(v):
            return f"{int(round(v)):,}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9)
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=8, leading=10)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
            bold_style = ParagraphStyle("BoldStyle", parent=cell_style, fontSize=8, bold=True)

            elems = []
            elems.append(Paragraph(titulo, title_style))
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))
            if ot_num:
                elems.append(Paragraph(f"<b>OT:</b> {ot_num}", subtitle_style))

            elems.append(Spacer(1, 3 * mm))

            table_data = [header]
            total_general = 0

            if not es_res_pe:
                current_fecha = None
                current_ot = None
                ot_subtotal = 0
                fecha_subtotal = 0

                def emit_subtotal_ot():
                    nonlocal ot_subtotal
                    if ot_subtotal > 0:
                        table_data.append([
                            Paragraph("", center_style),
                            Paragraph("<b>Subtotal OT</b>", bold_style),
                            Paragraph("", cell_style),
                            Paragraph("", center_style),
                            Paragraph(clt(ot_subtotal), right_style),
                        ])
                        ot_subtotal = 0

                def emit_subtotal_fecha():
                    nonlocal fecha_subtotal
                    if fecha_subtotal > 0:
                        table_data.append([
                            Paragraph("", center_style),
                            Paragraph("", cell_style),
                            Paragraph("<b>Subtotal Fecha</b>", bold_style),
                            Paragraph("", right_style),
                            Paragraph(clt(fecha_subtotal), right_style),
                        ])
                        fecha_subtotal = 0

                for m in qs.iterator():
                    numero = int(m.numero) if m.numero else 0
                    if numero == 0:
                        continue

                    fecha_key = m.fecha.strftime("%Y-%m-%d") if m.fecha else ""
                    proceso_val = int(m.proceso) if m.proceso else 0
                    enc_val = int(m.codencargado) if m.codencargado else 0
                    cantidad = m.cantidad if m.cantidad else 0

                    if current_fecha is not None and fecha_key != current_fecha:
                        emit_subtotal_ot()
                        emit_subtotal_fecha()

                    if current_ot is not None and numero != current_ot:
                        emit_subtotal_ot()

                    current_fecha = fecha_key
                    current_ot = numero

                    table_data.append([
                        Paragraph(str(numero), center_style),
                        Paragraph(procesos_map.get(proceso_val, ""), cell_style),
                        Paragraph(empleados_map.get(enc_val, ""), cell_style),
                        Paragraph(m.fecha.strftime("%d-%m-%Y") if m.fecha else "", center_style),
                        Paragraph(clq(cantidad), right_style),
                    ])
                    total_general += cantidad
                    ot_subtotal += cantidad
                    fecha_subtotal += cantidad

                emit_subtotal_ot()
                emit_subtotal_fecha()
            else:
                for m in qs.iterator():
                    numero = int(m.numero) if m.numero else 0
                    if numero == 0:
                        continue
                    
                    cantidad = m.cantidad if m.cantidad else 0
                    
                    table_data.append([
                        Paragraph(m.codigo.codigo if m.codigo else "", center_style),
                        Paragraph(m.codigo.descr if m.codigo else "", cell_style),
                        Paragraph(clq(cantidad), right_style),
                        Paragraph(m.codigo.um if m.codigo else "", center_style),
                    ])
                    total_general += cantidad

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            
            for i in range(1, len(table_data)):
                row_text = " ".join(str(c) for c in table_data[i])
                if "Subtotal" in row_text:
                    style_cmds.extend([
                        ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#e5e7eb")),
                        ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"),
                    ])
                else:
                    bg = colors.white if i % 2 == 1 else colors.HexColor("#f9fafb")
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
            
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)

            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[
                Paragraph(f'<b>Total General</b>  {clt(total_general)}', ParagraphStyle("Resumen", parent=styles["Normal"], fontSize=9, leading=12, alignment=1)),
            ]]
            resumen_tbl = Table(resumen_data, colWidths=[150 * mm], hAlign="RIGHT")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elems.append(resumen_tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=22 * mm, bottomMargin=18 * mm)
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = "res_pe.pdf" if es_res_pe else "resumen_ots.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    def _generar_pdf_detalle(self, request: HttpRequest) -> HttpResponse:
        ot_num = request.POST.get("ot", "").strip()
        usuario = str(request.user)

        try:
            ot_val = float(ot_num)
        except ValueError:
            return HttpResponse("OT inválida", content_type="text/plain")

        def get_data(tipo_cod, linea_filter):
            qs = Movs.objects.select_related("codigo").filter(
                tipo__cod__in=[tipo_cod],
                numero=ot_val
            )
            if linea_filter == "gt":
                qs = qs.filter(linea__gt=0)
            elif linea_filter == "ex0":
                qs = qs.exclude(linea=0)
            elif linea_filter == "eq1":
                qs = qs.filter(linea=1)

            rows = []
            for m in qs.iterator():
                cantidad = abs(m.cantidad) if m.cantidad else 0
                if cantidad == 0:
                    continue
                rows.append({
                    "codigo": m.codigo.codigo if m.codigo else "",
                    "nombre": m.codigo.descr if m.codigo else "",
                    "cantidad": cantidad,
                    "um": m.codigo.um if m.codigo else "",
                })
            return rows

        detalle_ot = get_data(8, "ex0")
        vale_consumo = get_data(10, "gt")
        parte_entrada = get_data(6, "gt")

        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}
        empleados_map = {e["cod"]: e["nombre"] for e in Empleados.objects.values("cod", "nombre")}

        ot_info = Movs.objects.filter(tipo=8, numero=ot_val, linea=0).first()
        proc_nom = procesos_map.get(int(ot_info.proceso), "") if ot_info and ot_info.proceso else ""
        enc_nom = empleados_map.get(int(ot_info.codencargado), "") if ot_info and ot_info.codencargado else ""
        fecha_str = ot_info.fecha.strftime("%d-%m-%Y") if ot_info and ot_info.fecha else ""

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/brand-logos/logo-home-grande.png")
        if not os.path.exists(logo_path):
            logo_path = None

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9)
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=8, leading=9)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
            bold_style = ParagraphStyle("BoldStyle", parent=cell_style, fontSize=8, bold=True)
            section_style = ParagraphStyle("SectionTitle", parent=styles["Heading3"], fontSize=10, bold=True, spaceAfter=3 * mm)

            elems = []
            elems.append(Paragraph(f"Detalle de Orden de Trabajo", title_style))
            elems.append(Paragraph(f"<b>OT:</b> {ot_num}", subtitle_style))
            elems.append(Paragraph(f"<b>Fecha:</b> {fecha_str}  |  <b>Encargado:</b> {enc_nom}  |  <b>Proceso:</b> {proc_nom}", subtitle_style))

            elems.append(Spacer(1, 4 * mm))

            header = ["Artículo", "Nombre", "Cantidad", "UM"]
            col_widths = [25 * mm, 70 * mm, 35 * mm, 20 * mm]

            def add_section(title, data):
                nonlocal elems
                elems.append(Paragraph(title, section_style))
                if not data:
                    elems.append(Paragraph("Sin datos", cell_style))
                    elems.append(Spacer(1, 3 * mm))
                    return
                
                table_data = [header]
                for row in data:
                    table_data.append([
                        Paragraph(str(row["codigo"]), center_style),
                        Paragraph(row["nombre"], cell_style),
                        Paragraph(clq(row["cantidad"]), right_style),
                        Paragraph(row["um"], center_style),
                    ])
                
                tbl = Table(table_data, colWidths=col_widths)
                style_cmds = [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
                for i in range(1, len(table_data)):
                    bg = colors.white if i % 2 == 1 else colors.HexColor("#f9fafb")
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
                tbl.setStyle(TableStyle(style_cmds))
                elems.append(tbl)
                elems.append(Spacer(1, 5 * mm))

            add_section("Detalle de Orden de Trabajo", detalle_ot)
            add_section("Vale de Consumo", vale_consumo)
            add_section("Parte de Entrada", parte_entrada)

            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=22 * mm, bottomMargin=18 * mm)
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="detalle_ot_{ot_num}.pdf"'
        return response