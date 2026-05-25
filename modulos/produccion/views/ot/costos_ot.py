from typing import Any
import io
import os
from datetime import datetime

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import connection
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.empleados import Empleados
from modulos.maestros.models.procesos import Procesos
from django.conf import settings


class IndexCostosOTView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/produccion/ot/costos_ot.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        if action == "listar_ots":
            return self._listar_ots()
        elif action == "listar_costos":
            return self._listar_costos(request)
        elif action == "calcular_costos":
            return self._calcular_costos(request)
        elif action == "generar_pdf":
            return self._generar_pdf(request)
        elif action == "generar_excel":
            return self._generar_excel(request)
        return JsonResponse({"success": False})

    def _calcular_costos(self, request: HttpRequest) -> JsonResponse:
        desde, hasta = self._get_ot_params(request)
        if desde is None or hasta is None:
            return JsonResponse({"success": False, "error": "Debe ingresar OT Desde y OT Hasta"})
        try:
            with connection.cursor() as cursor:
                cursor.execute("EXEC spCostosOT %s, %s", [desde, hasta])
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    def _listar_costos(self, request: HttpRequest) -> JsonResponse:
        desde, hasta = self._get_ot_params(request)
        qs = Movs.objects.select_related("codigo", "tipo").filter(
            tipo__cod=8
        ).exclude(numero__isnull=True).exclude(numero=0)
        if desde:
            qs = qs.filter(numero__gte=desde)
        if hasta:
            qs = qs.filter(numero__lte=hasta)
        qs = qs.order_by("numero", "linea")

        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}

        def fmt(v):
            if v is None:
                return "0"
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        data = []
        for m in qs.iterator():
            proc_val = int(m.proceso) if m.proceso else 0
            data.append({
                "ot": str(int(m.numero)) if m.numero else "",
                "fecha": m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                "proceso": procesos_map.get(proc_val, ""),
                "codigo": m.codigo.codigo if m.codigo else "",
                "articulo": m.codigo.descr if m.codigo else "",
                "cantidad": fmt(m.cantidad),
                "punit": fmt(m.punit),
                "neto": fmt(m.neto),
                "total": fmt(m.total),
            })
        return JsonResponse({"data": data})

    def _listar_ots(self) -> JsonResponse:
        qs = Movs.objects.filter(tipo__cod=8, linea=0).exclude(numero__isnull=True).exclude(numero=0)
        qs = qs.order_by("-numero").values("numero", "fecha", "proceso").distinct()
        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}
        ot_list = []
        for m in qs:
            numero = int(m["numero"]) if m["numero"] else 0
            if numero == 0:
                continue
            proc_val = int(m["proceso"]) if m["proceso"] else 0
            ot_list.append({
                "numero": numero,
                "fecha": m["fecha"].strftime("%d-%m-%Y") if m["fecha"] else "",
                "proceso": procesos_map.get(proc_val, ""),
            })
        return JsonResponse({"ots": ot_list})

    def _get_ot_params(self, request):
        ot_desde = request.POST.get("ot_desde", "").strip()
        ot_hasta = request.POST.get("ot_hasta", "").strip()
        desde = None
        hasta = None
        if ot_desde:
            try:
                desde = float(ot_desde)
            except ValueError:
                pass
        if ot_hasta:
            try:
                hasta = float(ot_hasta)
            except ValueError:
                pass
        return desde, hasta

    def _get_datos(self, request):
        desde, hasta = self._get_ot_params(request)
        qs = Movs.objects.select_related("codigo", "tipo").filter(
            tipo__cod=8, linea__gt=0
        ).exclude(numero__isnull=True).exclude(numero=0)
        if desde:
            qs = qs.filter(numero__gte=desde)
        if hasta:
            qs = qs.filter(numero__lte=hasta)
        return qs.order_by("numero", "linea")

    def _generar_pdf(self, request: HttpRequest) -> HttpResponse:
        desde, hasta = self._get_ot_params(request)
        usuario = str(request.user)
        qs = self._get_datos(request)

        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/brand-logos/logo-home-grande.png")
        if not os.path.exists(logo_path):
            logo_path = None

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9)
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=7, leading=9)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)

            elems = []
            elems.append(Paragraph("Costos por OT", title_style))
            rango = ""
            if desde:
                rango += f"desde OT {int(desde)}"
            if hasta:
                rango += f" hasta OT {int(hasta)}"
            if rango:
                elems.append(Paragraph(f"<b>Rango:</b> {rango}", subtitle_style))
            elems.append(Spacer(1, 3 * mm))

            header = ["OT", "Linea", "Código", "Artículo", "Cantidad", "PUnit", "Neto", "Total"]
            col_widths = [18 * mm, 12 * mm, 22 * mm, 50 * mm, 18 * mm, 18 * mm, 20 * mm, 22 * mm]
            table_data = [header]

            for m in qs.iterator():
                table_data.append([
                    Paragraph(str(int(m.numero)) if m.numero else "", center_style),
                    Paragraph(str(int(m.linea)) if m.linea else "", center_style),
                    Paragraph(m.codigo.codigo if m.codigo else "", cell_style),
                    Paragraph(m.codigo.descr if m.codigo else "", cell_style),
                    Paragraph(clq(m.cantidad) if m.cantidad else "0", right_style),
                    Paragraph(clq(m.punit) if m.punit else "0", right_style),
                    Paragraph(clq(m.neto) if m.neto else "0", right_style),
                    Paragraph(clq(m.total) if m.total else "0", right_style),
                ])

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
            for i in range(1, len(table_data)):
                bg = colors.white if i % 2 == 1 else colors.HexColor("#f9fafb")
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=22 * mm, bottomMargin=18 * mm)
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="costos_ot.pdf"'
        return response

    def _generar_excel(self, request: HttpRequest) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        qs = self._get_datos(request)

        wb = Workbook()
        ws = wb.active
        ws.title = "Costos OT"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["OT", "Linea", "Código", "Artículo", "Cantidad", "PUnit", "Neto", "Total"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        for m in qs.iterator():
            ws.append([
                int(m.numero) if m.numero else 0,
                int(m.linea) if m.linea else 0,
                m.codigo.codigo if m.codigo else "",
                m.codigo.descr if m.codigo else "",
                m.cantidad or 0,
                m.punit or 0,
                m.neto or 0,
                m.total or 0,
            ])

        row_num = ws.max_row
        for r in range(2, row_num + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.border = thin_border
                if col_idx in (5, 6, 7, 8):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.000'
                elif col_idx in (1, 2):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="costos_ot.xlsx"'
        return response


class ImprimirOTView(TemplateView):
    template_name = 'modulos/produccion/ot/imprimir.html'

    def get(self, request, *args, **kwargs):
        from django.conf import settings
        import os
        from datetime import datetime

        numero = float(kwargs.get('numero'))
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        encabezado = Movs.objects.filter(numero=numero, tipo__cod=8, linea=0).first()

        if not encabezado:
            return HttpResponse("OT no encontrada")

        detalles = list(Movs.objects.filter(
            numero=numero,
            tipo__cod=8
        ).exclude(linea=0).select_related('codigo').order_by('linea'))

        encargado_obj = Empleados.objects.filter(
            cod=encabezado.codencargado).first() if encabezado.codencargado else None
        proceso_obj = Procesos.objects.filter(cod=encabezado.proceso).first() if encabezado.proceso else None

        cell_style = ParagraphStyle("CellStyle", parent=getSampleStyleSheet()["Normal"], fontSize=7, leading=9)
        center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
        right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9)

            elems = []
            elems.append(Paragraph("Orden de Trabajo", title_style))
            elems.append(Paragraph(
                f"<b>N° OT:</b> {int(numero)} &nbsp;&nbsp;&nbsp;"
                f"<b>Fecha:</b> {encabezado.fecha.strftime('%d-%m-%Y') if encabezado.fecha else ''} &nbsp;&nbsp;&nbsp;"
                f"<b>Estado:</b> {encabezado.estado or ''}",
                subtitle_style,
            ))
            elems.append(Paragraph(
                f"<b>Encargado:</b> {encargado_obj.nombre if encargado_obj else ''} &nbsp;&nbsp;&nbsp;"
                f"<b>Proceso:</b> {proceso_obj.nombre if proceso_obj else ''}",
                subtitle_style,
            ))
            elems.append(Spacer(1, 3 * mm))

            header = ["DocRef", "Tipo", "Fecha", "Código", "Artículo", "Cant."]
            col_widths = [22 * mm, 18 * mm, 28 * mm, 30 * mm, 72 * mm, 20 * mm]
            table_data = [header]

            total_cant = 0

            for det in detalles:
                tipo_ref = int(det.tipodocref) if det.tipodocref else 0
                if tipo_ref == 7:
                    tipo_nombre = 'OR'
                elif tipo_ref == 6:
                    tipo_nombre = 'PE'
                else:
                    tipo_nombre = ''

                cant = abs(det.cantidad) if det.cantidad else 0
                total_cant += cant

                table_data.append([
                    Paragraph(str(int(det.docref)) if det.docref else '', center_style),
                    Paragraph(tipo_nombre, center_style),
                    Paragraph(det.fecha.strftime('%d-%m-%Y') if det.fecha else '', center_style),
                    Paragraph(det.codigo.codigo if det.codigo else '', cell_style),
                    Paragraph(det.codigo.descr if det.codigo else '', cell_style),
                    Paragraph(str(int(cant)), right_style),
                ])

            if len(detalles) > 0:
                table_data.append([
                    Paragraph("", cell_style),
                    Paragraph("", cell_style),
                    Paragraph("", cell_style),
                    Paragraph("", cell_style),
                    Paragraph("TOTAL", cell_style),
                    Paragraph(str(total_cant), right_style),
                ])

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]
            if len(detalles) > 0:
                style_cmds.append(
                    ("BACKGROUND", (0, len(detalles) + 1), (-1, len(detalles) + 1), colors.HexColor("#e5e7eb")))
                style_cmds.append(("FONTNAME", (0, len(detalles) + 1), (-1, len(detalles) + 1), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)
            elems.append(Spacer(1, 10 * mm))

            elems.append(Paragraph("Entrega de Trabajo", title_style))
            elems.append(Spacer(1, 3 * mm))

            ent_headers = ["Código Terminado", "Cant.", "Fecha", "Código Insumo", "Cant.", "Fecha"]
            ent_widths = [22 * mm, 18 * mm, 28 * mm, 30 * mm, 72 * mm, 20 * mm]
            ent_data = [ent_headers]

            for _ in range(10):
                ent_data.append(['', '', '', '', '', ''])

            t_ent = Table(ent_data, colWidths=ent_widths, repeatRows=1)
            t_ent.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
            elems.append(t_ent)
            elems.append(Spacer(1, 5 * mm))

            elems.append(Paragraph("Fecha de Entrega: _________________", subtitle_style))
            elems.append(Spacer(1, 15 * mm))

            datos_firmas = [['Encargado', 'Supervisor']]
            t_fir = Table(datos_firmas, colWidths=[80 * mm, 80 * mm])
            t_fir.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 30),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("LINEABOVE", (0, 0), (0, 0), 0.5, colors.HexColor("#d1d5db")),
                ("LINEABOVE", (1, 0), (1, 0), 0.5, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
            elems.append(t_fir)

            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm,
                                               preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm,
                                               preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=22 * mm, bottomMargin=18 * mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        buf.seek(0)
        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="OT_{int(numero)}.pdf"'
        return response
