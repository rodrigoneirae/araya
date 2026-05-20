import io
import os
from datetime import datetime
from typing import Any

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.empleados import Empleados
from modulos.maestros.models.docs import Docs
from modulos.maestros.models.procesos import Procesos


class IndexInformeProduccionProcesoView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/produccion/informes/produccion_proceso.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        if action == "listar_procesos":
            return self._listar_procesos()
        elif action == "generar_pdf":
            return self._generar_pdf(request)
        elif action == "generar_excel":
            return self._generar_excel(request)
        return JsonResponse({"success": False})

    def _listar_procesos(self) -> JsonResponse:
        procesos = Procesos.objects.values("cod", "nombre").order_by("nombre")
        return JsonResponse({"procesos": list(procesos)})

    def _get_datos(self, fecha_inicio: str, fecha_fin: str, proceso_cod: str | None):
        fi = None
        ff = None
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d")
            except Exception:
                pass
        if fecha_fin:
            try:
                ff = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d")
            except Exception:
                pass

        qs = Movs.objects.select_related("codigo", "tipo").filter(tipo__cod__in=[6], linea__gt=0)

        if fi:
            qs = qs.filter(fecha__gte=fi)
        if ff:
            qs = qs.filter(fecha__lte=ff.replace(hour=23, minute=59, second=59))
        if proceso_cod:
            try:
                proc_val = int(proceso_cod)
                qs = qs.filter(proceso=proc_val)
            except ValueError:
                pass
        else:
            qs = qs.exclude(proceso__isnull=True).exclude(proceso=0)

        qs = qs.order_by("proceso", "codigo", "fecha", "numero")

        empleados_map = {e["cod"]: e["nombre"] for e in Empleados.objects.values("cod", "nombre")}
        docs_map = {d["cod"]: d["nombre"] for d in Docs.objects.values("cod", "nombre")}
        procesos_map = {p["cod"]: p["nombre"] for p in Procesos.objects.values("cod", "nombre")}

        data = {}
        for m in qs.iterator():
            proceso_val = int(m.proceso) if m.proceso else 0
            proc_nom = procesos_map.get(proceso_val, f"Proceso {proceso_val}")
            
            if proceso_val not in data:
                data[proceso_val] = {"nombre": proc_nom, "articulos": {}}

            cod_art = m.codigo.codigo if m.codigo else ""
            nom_art = m.codigo.descr if m.codigo else ""
            um_art = m.codigo.um if m.codigo else ""
            numero_ot = m.numero if m.numero else ""
            fecha = m.fecha.strftime("%d-%m-%Y") if m.fecha else ""
            tipo_doc = docs_map.get(int(m.tipo.cod), "") if m.tipo else ""
            
            enc_cod = int(m.codencargado) if m.codencargado else 0
            enc_nom = empleados_map.get(enc_cod, "")

            if cod_art not in data[proceso_val]["articulos"]:
                data[proceso_val]["articulos"][cod_art] = {"nombre": nom_art, "um": um_art, "movimientos": []}

            cantidad = abs(m.cantidad) if m.cantidad else 0
            if cantidad == 0:
                continue

            data[proceso_val]["articulos"][cod_art]["movimientos"].append({
                "fecha": fecha,
                "ot": numero_ot,
                "tipo": tipo_doc,
                "cantidad": cantidad,
                "um": um_art,
                "encargado": enc_nom,
            })

        return data

    def _generar_excel(self, request: HttpRequest) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        fecha_inicio = request.POST.get("fecha_inicio", "").strip()
        fecha_fin = request.POST.get("fecha_fin", "").strip()
        proceso_cod = request.POST.get("proceso", "").strip()

        data = self._get_datos(fecha_inicio, fecha_fin, proceso_cod if proceso_cod else None)

        wb = Workbook()
        ws = wb.active
        ws.title = "Producción por Proceso"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Artículo", "Nombre", "Fecha", "N° OT", "Tipo", "Cant", "UM", "Encargado"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def agregar_filas(data_dict):
            if proceso_cod:
                enc_data = list(data_dict.values())[0] if data_dict else None
                if enc_data:
                    for art_cod, art_data in sorted(enc_data["articulos"].items()):
                        for mov in art_data["movimientos"]:
                            ws.append([art_cod, art_data["nombre"], mov["fecha"], mov["ot"], mov["tipo"], mov["cantidad"], mov["um"], mov["encargado"]])
            else:
                for proc_cod, proc_data in sorted(data_dict.items()):
                    for art_cod, art_data in sorted(proc_data["articulos"].items()):
                        for mov in art_data["movimientos"]:
                            ws.append([art_cod, art_data["nombre"], mov["fecha"], mov["ot"], mov["tipo"], mov["cantidad"], mov["um"], mov["encargado"]])

            row_num = ws.max_row
            for r in range(2, row_num + 1):
                for col_idx in range(1, 9):
                    cell = ws.cell(row=r, column=col_idx)
                    cell.border = thin_border
                    if col_idx == 6:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.000'
                    elif col_idx in [1, 2, 8]:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        agregar_filas(data)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "produccion_proceso.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf(self, request: HttpRequest) -> HttpResponse:
        fecha_inicio = request.POST.get("fecha_inicio", "").strip()
        fecha_fin = request.POST.get("fecha_fin", "").strip()
        proceso_cod = request.POST.get("proceso", "").strip()
        usuario = str(request.user)

        data = self._get_datos(fecha_inicio, fecha_fin, proceso_cod if proceso_cod else None)

        fi = None
        ff = None
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d")
            except Exception:
                pass
        if fecha_fin:
            try:
                ff = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d")
            except Exception:
                pass

        fecha_inf = ""
        if fi or ff:
            fecha_inf = "del "
            if fi:
                fecha_inf += fi.strftime("%d-%m-%Y")
            fecha_inf += " al "
            if ff:
                fecha_inf += ff.strftime("%d-%m-%Y")

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/brand-logos/logo-home-grande.png")
        if not os.path.exists(logo_path):
            logo_path = None

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")
        
        def clt(v):
            return f"{int(round(v)):,}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9)
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=7, leading=9)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
            bold_style = ParagraphStyle("BoldStyle", parent=cell_style, fontSize=7, bold=True)

            elems = []
            elems.append(Paragraph("Informe de Producción por Proceso", title_style))
            proc_nombre = "Todos los Procesos" if not proceso_cod else (list(data.values())[0]['nombre'] if data else '')
            elems.append(Paragraph(f"<b>Proceso:</b> {proc_nombre}", subtitle_style))
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 3 * mm))

            header = ["Artículo", "Nombre", "Fecha", "N° OT", "Tipo", "Cant", "UM", "Encargado"]
            col_widths = [20 * mm, 38 * mm, 18 * mm, 15 * mm, 22 * mm, 16 * mm, 10 * mm, 25 * mm]

            if not data:
                return HttpResponse("No hay datos", content_type="text/plain")

            table_data = [header]
            total_general = 0
            proc_subtotal = 0

            def add_proc_subtotal():
                nonlocal proc_subtotal
                if proc_subtotal > 0:
                    table_data.append([
                        Paragraph("", center_style),
                        Paragraph("<b>Total Proceso</b>", bold_style),
                        Paragraph("", center_style),
                        Paragraph("", center_style),
                        Paragraph("", cell_style),
                        Paragraph(clt(proc_subtotal), ParagraphStyle("RightBold", parent=right_style, fontSize=7, bold=True)),
                        Paragraph("", center_style),
                        Paragraph("", cell_style),
                    ])

            if proceso_cod:
                proc_data = list(data.values())[0] if data else None
                if proc_data:
                    for art_cod, art_data in sorted(proc_data["articulos"].items()):
                        if not art_data["movimientos"]:
                            continue
                        for mov in art_data["movimientos"]:
                            table_data.append([
                                Paragraph(str(art_cod), center_style),
                                Paragraph(art_data["nombre"], cell_style),
                                Paragraph(mov["fecha"], center_style),
                                Paragraph(str(mov["ot"]) if mov["ot"] else "", center_style),
                                Paragraph(mov["tipo"] or "", cell_style),
                                Paragraph(clq(mov["cantidad"]), right_style),
                                Paragraph(mov["um"], center_style),
                                Paragraph(mov["encargado"] or "", cell_style),
                            ])
                            total_general += mov["cantidad"]
                            proc_subtotal += mov["cantidad"]
                add_proc_subtotal()
            else:
                for proc_cod_val, proc_data in sorted(data.items()):
                    proc_subtotal = 0
                    
                    table_data.append([
                        Paragraph("", center_style),
                        Paragraph(f"<b>=== {proc_data['nombre']} ===</b>", bold_style),
                        Paragraph("", center_style),
                        Paragraph("", center_style),
                        Paragraph("", cell_style),
                        Paragraph("", right_style),
                        Paragraph("", center_style),
                        Paragraph("", cell_style),
                    ])
                    
                    for art_cod, art_data in sorted(proc_data["articulos"].items()):
                        if not art_data["movimientos"]:
                            continue
                        for mov in art_data["movimientos"]:
                            table_data.append([
                                Paragraph(str(art_cod), center_style),
                                Paragraph(art_data["nombre"], cell_style),
                                Paragraph(mov["fecha"], center_style),
                                Paragraph(str(mov["ot"]) if mov["ot"] else "", center_style),
                                Paragraph(mov["tipo"] or "", cell_style),
                                Paragraph(clq(mov["cantidad"]), right_style),
                                Paragraph(mov["um"], center_style),
                                Paragraph(mov["encargado"] or "", cell_style),
                            ])
                            total_general += mov["cantidad"]
                            proc_subtotal += mov["cantidad"]
                    
                    add_proc_subtotal()

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("ALIGN", (2, 0), (4, -1), "CENTER"),
                ("ALIGN", (6, 0), (6, -1), "CENTER"),
            ]
            
            for i in range(1, len(table_data)):
                row_text = str(table_data[i][1]) if len(table_data[i]) > 1 else ""
                if "===" in row_text or "Total Proceso" in row_text or "Total General" in row_text:
                    style_cmds.extend([
                        ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#e5e7eb")),
                        ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"),
                    ])
                else:
                    bg = colors.white if i % 2 == 1 else colors.HexColor("#f9fafb")
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
            
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)

            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[
                Paragraph(f'<b>Total General</b>  {clt(total_general)}', ParagraphStyle("Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1)),
            ]]
            resumen_tbl = Table(resumen_data, colWidths=[200 * mm], hAlign="RIGHT")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elems.append(resumen_tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=22 * mm, bottomMargin=18 * mm)
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = "produccion_proceso.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response