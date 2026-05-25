import io
import os
from typing import Any
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.articulos import Articulos
from modulos.maestros.models.prov_cliente import Provclientes


class IndexInformeCompraProvedoresView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/inventario/informes/compra_provedores.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        handlers = {
            "listar_proveedores": lambda _: self._listar_proveedores(),
            "buscar_proveedor": lambda d: self._buscar_proveedor(d.get("rut")),
            "info_informe": lambda d: self._info_informe(d),
            "info_libro": lambda d: self._info_libro(d),
            "generar_pdf_informe": lambda d: self._generar_pdf_informe(request, d),
            "generar_pdf_libro": lambda d: self._generar_pdf_libro(request, d),
            "generar_excel_informe": lambda d: self._generar_excel_informe(request, d),
            "generar_excel_libro": lambda d: self._generar_excel_libro(request, d),
        }
        handler = handlers.get(action, lambda _: JsonResponse({"success": False}))
        return handler(request.POST)

    def _parse_fechas(self, data: dict) -> tuple:
        fi = None
        fc = None
        fecha_inicio = data.get("fecha_inicio", "").strip()
        fecha_corte = data.get("fecha_corte", "").strip()
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            except Exception:
                pass
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte, "%Y-%m-%d")
            except Exception:
                pass
        return fi, fc

    def _listar_proveedores(self) -> JsonResponse:
        proveedores = (
            Provclientes.objects
            .exclude(rut='')
            .values("rut", "nombre")
            .order_by("nombre")
        )
        return JsonResponse({"proveedores": list(proveedores)})

    def _buscar_proveedor(self, rut: str | None) -> JsonResponse:
        if not rut:
            return JsonResponse({"success": False})
        try:
            prov = Provclientes.objects.get(rut=rut.strip())
            return JsonResponse({
                "success": True,
                "data": {"rut": prov.rut, "nombre": prov.nombre},
            })
        except Provclientes.DoesNotExist:
            return JsonResponse({"success": False, "message": "Proveedor no encontrado"})

    def _get_proveedores_map(self) -> dict:
        return {p["rut"]: p["nombre"] for p in Provclientes.objects.values("rut", "nombre")}

    def _base_qs_informe(self, rut: str, fi, fc):
        qs = (
            Movs.objects
            .select_related("codigo")
            .filter(tipo=7, linea__gt=0)
        )
        if rut:
            qs = qs.filter(rut=rut)
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))
        return qs.order_by("rut", "codigo")

    def _agrupar_por_proveedor(self, qs, proveedores_map: dict) -> list[dict]:
        movimientos = []
        rut_actual = None
        sub_cantidad = 0
        sub_monto = 0

        def _flush():
            nonlocal sub_cantidad, sub_monto
            movimientos.append({
                "_subtotal": True,
                "_rut": rut_actual,
                "_proveedor": proveedores_map.get(rut_actual, ""),
                "_cantidad": sub_cantidad,
                "_monto": sub_monto,
            })
            sub_cantidad = 0
            sub_monto = 0

        first_in_group = True
        for m in qs.iterator():
            m_rut = m.rut or ""
            if m_rut != rut_actual:
                if rut_actual is not None:
                    _flush()
                rut_actual = m_rut
                first_in_group = True

            cantidad = m.cantidad or 0
            monto = cantidad * (m.punit or 0)
            sub_cantidad += cantidad
            sub_monto += monto

            movimientos.append({
                "_subtotal": False,
                "_first_in_group": first_in_group,
                "articulo_codigo": m.codigo.codigo if m.codigo else "",
                "articulo_nombre": m.codigo.descr if m.codigo else "",
                "fecha": m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                "rut": m_rut,
                "proveedor_nombre": proveedores_map.get(m_rut, ""),
                "cantidad": cantidad,
                "punit": m.punit or 0,
                "total": monto,
                "um": m.codigo.um if m.codigo else "",
            })
            first_in_group = False

        if rut_actual is not None:
            _flush()

        total_cant = sum(m["_cantidad"] for m in movimientos if m.get("_subtotal"))
        total_monto = sum(m["_monto"] for m in movimientos if m.get("_subtotal"))
        for m in movimientos:
            if m.get("_subtotal"):
                m["_participacion_cant"] = round(m["_cantidad"] / total_cant * 100, 2) if total_cant else 0
                m["_participacion_monto"] = round(m["_monto"] / total_monto * 100, 2) if total_monto else 0

        return movimientos

    def _info_informe(self, data: dict) -> JsonResponse:
        rut = data.get("rut", "").strip()
        fi, fc = self._parse_fechas(data)
        qs = self._base_qs_informe(rut, fi, fc)
        proveedores_map = self._get_proveedores_map()
        movimientos = self._agrupar_por_proveedor(qs, proveedores_map)
        return JsonResponse({"success": True, "data": movimientos})

    def _agrupar_libro(self, rut: str, fi, fc) -> list[dict]:
        qs = (
            Movs.objects
            .select_related("tipo")
            .filter(tipo__in=[7, 11], linea__gt=0)
        )
        if rut:
            qs = qs.filter(rut=rut)
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        proveedores_map = self._get_proveedores_map()

        docs = {}
        for m in qs.iterator():
            doc_key = (m.tipo_id, m.numero or 0, m.rut or "", m.fecha, m.docref or 0)
            if doc_key not in docs:
                docs[doc_key] = {
                    "tipo_nombre": m.tipo.nombre if m.tipo else "",
                    "numero": str(int(m.numero)) if m.numero else "",
                    "fecha": m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                    "rut": m.rut or "",
                    "proveedor_nombre": proveedores_map.get(m.rut, ""),
                    "docref": str(int(m.docref)) if m.docref else "",
                    "cantidad": 0,
                    "neto": 0,
                }
            cantidad = m.cantidad or 0
            docs[doc_key]["cantidad"] += cantidad
            docs[doc_key]["neto"] += cantidad * (m.punit or 0)

        return sorted(docs.values(), key=lambda d: (d["tipo_nombre"], d["numero"]))

    def _info_libro(self, data: dict) -> JsonResponse:
        rut = data.get("rut", "").strip()
        fi, fc = self._parse_fechas(data)
        movimientos = self._agrupar_libro(rut, fi, fc)
        return JsonResponse({"success": True, "data": movimientos})

    # --- PDF generation helpers ---
    def _build_pdf(
        self, request: HttpRequest, data: dict, title: str,
        table_headers: list[str], col_widths: list[float],
        row_builder, is_libro: bool = False,
    ) -> HttpResponse:
        rut = data.get("rut", "").strip()
        fi, fc = self._parse_fechas(data)
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        proveedores_map = self._get_proveedores_map()

        if is_libro:
            qs = (
                Movs.objects.select_related("codigo", "tipo")
                .filter(tipo__in=[7, 11], linea__gt=0)
            )
            if rut:
                qs = qs.filter(rut=rut)
            if fi:
                qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
            if fc:
                qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))
            qs = qs.order_by("tipo", "numero", "codigo")
        else:
            qs = self._base_qs_informe(rut, fi, fc)

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        prov_nombre = ""
        if rut:
            try:
                prov = Provclientes.objects.get(rut=rut)
                prov_nombre = prov.nombre
            except Provclientes.DoesNotExist:
                pass

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=2 * mm, fontSize=8,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=5.5, leading=8,
            )
            right_style = ParagraphStyle(
                "RightStyle", parent=cell_style, alignment=2,
            )
            center_style = ParagraphStyle(
                "CenterStyle", parent=cell_style, alignment=1,
            )
            resumen_style = ParagraphStyle(
                "Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1,
            )

            total_cant_pdf = 0
            total_monto_pdf = 0
            elems = []
            elems.append(Paragraph(title, title_style))
            if rut:
                elems.append(Paragraph(
                    f"<b>RUT:</b> {rut} &nbsp;&nbsp;&nbsp;"
                    f"<b>Nombre:</b> {prov_nombre}",
                    subtitle_style,
                ))
            else:
                elems.append(Paragraph("<b>Todos los proveedores</b>", subtitle_style))

            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 2 * mm))

            table_data = [table_headers]
            subtotal_rows = []

            if is_libro:
                for m in qs.iterator():
                    cantidad = m.cantidad or 0
                    monto = cantidad * (m.punit or 0)
                    row = row_builder(m, cantidad, monto, proveedores_map, cell_style, center_style, right_style)
                    table_data.append(row)
            else:
                sub_style = ParagraphStyle(
                    "Subtotal", parent=getSampleStyleSheet()["Normal"],
                    fontSize=6, leading=9, fontName="Helvetica-Bold",
                    textColor=colors.HexColor("#4b5563"),
                )
                sub_right = ParagraphStyle("SubR", parent=sub_style, alignment=2)

                rows_buffer = []
                subtotal_info = []
                rut_actual = None
                sub_cantidad = 0
                sub_monto = 0

                for m in qs.iterator():
                    m_rut = m.rut or ""
                    if m_rut != rut_actual:
                        if rut_actual is not None:
                            subtotal_info.append((rut_actual, proveedores_map.get(rut_actual, ""), sub_cantidad, sub_monto))
                            sub_cantidad = 0
                            sub_monto = 0
                        rut_actual = m_rut

                    cantidad = m.cantidad or 0
                    monto = cantidad * (m.punit or 0)
                    sub_cantidad += cantidad
                    sub_monto += monto

                    row = row_builder(m, cantidad, monto, proveedores_map, cell_style, center_style, right_style)
                    rows_buffer.append((m_rut, row))

                if rut_actual is not None:
                    subtotal_info.append((rut_actual, proveedores_map.get(rut_actual, ""), sub_cantidad, sub_monto))

                total_cant_pdf = sum(s[2] for s in subtotal_info)
                total_monto_pdf = sum(s[3] for s in subtotal_info)
                subtotal_idx = 0
                rut_actual = None

                for m_rut, row in rows_buffer:
                    if m_rut != rut_actual:
                        if rut_actual is not None:
                            _rut_st, _prov_st, _cant_st, _monto_st = subtotal_info[subtotal_idx]
                            _part = round(_monto_st / total_monto_pdf * 100, 2) if total_monto_pdf else 0
                            table_data.append(row_builder(
                                None, _cant_st, _monto_st, None,
                                sub_style, sub_right, sub_right, is_subtotal=True,
                                subtotal_rut=_rut_st, subtotal_prov=_prov_st,
                                participacion=_part,
                            ))
                            subtotal_rows.append(len(table_data) - 1)
                            subtotal_idx += 1
                        rut_actual = m_rut
                    table_data.append(row)

                if subtotal_idx < len(subtotal_info):
                    _rut_st, _prov_st, _cant_st, _monto_st = subtotal_info[subtotal_idx]
                    _part = round(_monto_st / total_monto_pdf * 100, 2) if total_monto_pdf else 0
                    table_data.append(row_builder(
                        None, _cant_st, _monto_st, None,
                        sub_style, sub_right, sub_right, is_subtotal=True,
                        subtotal_rut=_rut_st, subtotal_prov=_prov_st,
                        participacion=_part,
                    ))
                    subtotal_rows.append(len(table_data) - 1)

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]
            if not is_libro:
                for sr in subtotal_rows:
                    style_cmds.append(("BACKGROUND", (0, sr), (-1, sr), colors.HexColor("#e5e7eb")))
                    style_cmds.append(("FONTNAME", (0, sr), (-1, sr), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)

            if not is_libro:
                elems.append(Spacer(1, 5 * mm))
                resumen_data = [[
                    Paragraph(f'<b>Total Cantidad</b><br/><font color="#2563eb">{clq(total_cant_pdf)}</font>', resumen_style),
                    Paragraph(f'<b>Monto Total</b><br/><font color="#7c3aed">{cl(total_monto_pdf)}</font>', resumen_style),
                    Paragraph(f'<b>Total Registros</b><br/><font color="#4b5563">{len(table_data) - 1 - len(subtotal_rows)}</font>', resumen_style),
                ]]
                resumen_tbl = Table(resumen_data, colWidths=[42*mm, 42*mm, 42*mm], hAlign="CENTER")
                resumen_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]))
                elems.append(resumen_tbl)

            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w/2, 10*mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w-15*mm, 10*mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=12*mm, rightMargin=12*mm,
            topMargin=18*mm, bottomMargin=16*mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()
        filename_suffix = "libro" if is_libro else "informe"
        filename = f"compra_proveedores_{filename_suffix}_{rut or 'todos'}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    def _generar_pdf_informe(self, request: HttpRequest, data: dict) -> HttpResponse:
        headers = ["RUT", "Proveedor", "Artículo", "Fecha", "Cant.", "Part.", "P.Unit", "Total"]
        col_widths = [20*mm, 36*mm, 36*mm, 15*mm, 15*mm, 14*mm, 16*mm, 22*mm]

        def row_builder(m, cantidad, monto, prov_map, c_style, center_style, r_style,
                        is_subtotal=False, subtotal_rut=None, subtotal_prov=None,
                        participacion=None):
            if is_subtotal:
                return [
                    Paragraph(f"<b>Subtotal {subtotal_rut}</b>", c_style),
                    Paragraph(f"<b>{subtotal_prov}</b>", c_style),
                    Paragraph("", c_style), Paragraph("", c_style),
                    Paragraph(f"<b>{self._clq(cantidad)}</b>", r_style),
                    Paragraph(f"<b>{(participacion or 0):.2f}%</b>", r_style),
                    Paragraph("", c_style),
                    Paragraph(f"<b>{self._cl(monto)}</b>", r_style),
                ]
            art_label = f"{m.codigo.codigo if m.codigo else ''} {m.codigo.descr if m.codigo else ''}".strip()
            return [
                Paragraph(m.rut or "", center_style),
                Paragraph(prov_map.get(m.rut, ""), c_style),
                Paragraph(art_label, c_style),
                Paragraph(m.fecha.strftime("%d-%m-%Y") if m.fecha else "", center_style),
                Paragraph(f"<b>{self._clq(cantidad)}</b>", r_style),
                Paragraph("", c_style),
                Paragraph(self._cl(m.punit or 0), r_style),
                Paragraph(self._cl(monto), r_style),
            ]

        self._cl = lambda v: f"{int(round(float(v))):,}".replace(",", ".")
        self._clq = lambda v: f"{int(round(float(v))):,}".replace(",", ".") if float(v) == int(float(v)) else f"{float(v):,.3f}".replace(",", ".")

        return self._build_pdf(
            request, data, "Compra por Proveedor - Informe",
            headers, col_widths, row_builder, is_libro=False,
        )

    def _generar_pdf_libro(self, request: HttpRequest, data: dict) -> HttpResponse:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        rut = data.get("rut", "").strip()
        fi, fc = self._parse_fechas(data)
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        docs = self._agrupar_libro(rut, fi, fc)

        logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/brand-logos/logo-home-grande.png")
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4*mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=2*mm, fontSize=8)
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=7, leading=10)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)

            elems = []
            elems.append(Paragraph("Libro de Compras", title_style))

            prov_nombre = ""
            if rut:
                try:
                    prov = Provclientes.objects.get(rut=rut)
                    prov_nombre = prov.nombre
                except Provclientes.DoesNotExist:
                    pass
                elems.append(Paragraph(f"<b>RUT:</b> {rut} &nbsp;&nbsp;&nbsp;<b>Nombre:</b> {prov_nombre}", subtitle_style))
            else:
                elems.append(Paragraph("<b>Todos los proveedores</b>", subtitle_style))

            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))
            elems.append(Spacer(1, 2*mm))

            headers = ["Tipo", "Número", "Cant.", "Neto", "Proveedor", "Fecha", "RUT", "DocRef"]
            col_widths = [20*mm, 18*mm, 18*mm, 22*mm, 36*mm, 18*mm, 20*mm, 16*mm]
            table_data = [headers]

            for d in docs:
                table_data.append([
                    Paragraph(d["tipo_nombre"], center_style),
                    Paragraph(d["numero"], center_style),
                    Paragraph(clq(d["cantidad"]), right_style),
                    Paragraph(cl(d["neto"]), right_style),
                    Paragraph(d["proveedor_nombre"], cell_style),
                    Paragraph(d["fecha"], center_style),
                    Paragraph(d["rut"], center_style),
                    Paragraph(d["docref"], center_style),
                ])

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]))
            elems.append(tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w/2, 10*mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w-15*mm, 10*mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=18*mm, bottomMargin=16*mm)
        doc.build(build_elements(), canvasmaker=_Canvas)
        pdf_bytes = buf.getvalue()
        buf.close()
        filename = f"compra_proveedores_libro_{rut or 'todos'}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    # --- Excel generation ---
    def _generar_excel_informe(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        rut = data.get("rut", "").strip()
        fi, fc = self._parse_fechas(data)
        qs = self._base_qs_informe(rut, fi, fc)
        proveedores_map = self._get_proveedores_map()
        movimientos = self._agrupar_por_proveedor(qs, proveedores_map)

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["RUT", "Proveedor", "Artículo", "Fecha", "Cantidad", "Part.", "P.Unit", "Total"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        sub_font = Font(name="Calibri", bold=True, size=10)
        sub_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        for item in movimientos:
            if item.get("_subtotal"):
                ws.append([
                    f"Subtotal {item['_rut']}", item["_proveedor"], "", "",
                    item["_cantidad"], f"{item.get('_participacion_monto', 0):.2f}%",
                    "", item["_monto"],
                ])
                row_num = ws.max_row
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = sub_font
                    cell.fill = sub_fill
                    cell.border = thin_border
                    cell.alignment = center_align
            else:
                art_label = f"{item.get('articulo_codigo', '')} {item.get('articulo_nombre', '')}".strip()
                ws.append([
                    item["rut"] if item.get("_first_in_group") else "",
                    item["proveedor_nombre"] if item.get("_first_in_group") else "",
                    art_label, item["fecha"], item["cantidad"],
                    "",
                    item["punit"], item["total"],
                ])
                row_num = ws.max_row
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = center_align

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"compra_proveedores_informe_{rut or 'todos'}.xlsx"
        response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_excel_libro(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        rut = data.get("rut", "").strip()
        fi, fc = self._parse_fechas(data)
        docs = self._agrupar_libro(rut, fi, fc)

        wb = Workbook()
        ws = wb.active
        ws.title = "Libro de Compras"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["Tipo", "Número", "Cantidad", "Neto", "Proveedor", "Fecha", "RUT", "DocRef"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center")

        for d in docs:
            ws.append([
                d["tipo_nombre"],
                d["numero"],
                d["cantidad"],
                d["neto"],
                d["proveedor_nombre"],
                d["fecha"],
                d["rut"],
                d["docref"],
            ])
            row_num = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_align

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"compra_proveedores_libro_{rut or 'todos'}.xlsx"
        response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
