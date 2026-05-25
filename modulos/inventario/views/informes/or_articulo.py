import io
import os
from typing import Any
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.articulos import Articulos
from modulos.maestros.models.bodegas import Bodegas
from modulos.maestros.models.prov_cliente import Provclientes


class IndexInformeOrArticuloView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/inventario/informes/or_articulo.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        handlers = {
            "listar_articulos": lambda _: self._listar_articulos(),
            "buscar_articulo": lambda d: self._buscar_articulo(d.get("codigo")),
            "info_proveedor": lambda d: self._info_proveedor(d),
            "informe_mensual": lambda d: self._informe_mensual(d),
            "generar_pdf_info": lambda d: self._generar_pdf_info(request, d),
            "generar_excel_info": lambda d: self._generar_excel_info(request, d),
            "generar_pdf_mensual": lambda d: self._generar_pdf_mensual(request, d),
            "generar_excel_mensual": lambda d: self._generar_excel_mensual(request, d),
        }
        handler = handlers.get(action, lambda _: JsonResponse({"success": False}))
        return handler(request.POST)

    def _listar_articulos(self) -> JsonResponse:
        articulos = (
            Articulos.objects
            .exclude(codigo='')
            .values("codigo", "descr", "um")
            .order_by("descr")
        )
        return JsonResponse({"articulos": list(articulos)})

    def _buscar_articulo(self, codigo: str | None) -> JsonResponse:
        if not codigo:
            return JsonResponse({"success": False})
        try:
            art = Articulos.objects.get(codigo=codigo.strip())
            return JsonResponse({
                "success": True,
                "data": {
                    "codigo": art.codigo,
                    "descr": art.descr,
                    "um": art.um,
                }
            })
        except Articulos.DoesNotExist:
            return JsonResponse({"success": False, "message": "Artículo no encontrado"})

    def _parse_fechas(self, data: dict) -> tuple:
        fi = None
        fc = None
        fecha_inicio = data.get("fecha_inicio", "").strip()
        fecha_corte = data.get("fecha_corte", "").strip()
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            except Exception:
                pass
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte, "%Y-%m-%d")
            except Exception:
                pass
        return fi, fc

    def _agrupar_por_rut(self, qs, bodegas_map: dict, proveedores_map: dict) -> list[dict]:
        movimientos = []
        rut_actual = None
        sub_cant_entradas = 0
        sub_cant_salidas = 0
        sub_monto = 0

        def _flush_subtotal():
            nonlocal sub_cant_entradas, sub_cant_salidas, sub_monto
            saldo = sub_cant_entradas - sub_cant_salidas
            movimientos.append({
                "_subtotal": True,
                "_rut": rut_actual,
                "_entradas": sub_cant_entradas,
                "_salidas": sub_cant_salidas,
                "_saldo": saldo,
                "_monto": sub_monto,
            })
            sub_cant_entradas = 0
            sub_cant_salidas = 0
            sub_monto = 0

        for m in qs.iterator():
            if not m.rut:
                continue
            if m.rut != rut_actual:
                if rut_actual is not None:
                    _flush_subtotal()
                rut_actual = m.rut

            signo = m.tipo.signo if m.tipo else 0
            cantidad = m.cantidad or 0
            if signo > 0:
                sub_cant_entradas += cantidad
            if signo < 0:
                sub_cant_salidas += abs(cantidad)
            monto = cantidad * (m.punit or 0)
            sub_monto += monto

            prov_nombre = proveedores_map.get(m.rut, "")

            movimientos.append({
                "_subtotal": False,
                "fecha": m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                "rut": m.rut,
                "proveedor_nombre": prov_nombre,
                "cantidad": cantidad,
                "punit": m.punit or 0,
                "total": monto,
                "tipo_nombre": m.tipo.nombre if m.tipo else "",
                "numero": str(int(m.numero)) if m.numero else "",
                "bodega_nombre": bodegas_map.get(m.bodega, str(m.bodega or "")),
                "signo": signo,
            })

        if rut_actual is not None:
            _flush_subtotal()

        return movimientos

    def _info_proveedor(self, data: dict) -> JsonResponse:
        codigo = data.get("codigo", "").strip()
        if not codigo:
            return JsonResponse({"success": False, "message": "Código de artículo requerido"})
        fi, fc = self._parse_fechas(data)

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(codigo=codigo, tipo__isnull=False)
        )
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("rut", "fecha")

        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}
        proveedores_map = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_map[p["rut"]] = p["nombre"]

        movimientos = self._agrupar_por_rut(qs, bodegas_map, proveedores_map)

        return JsonResponse({"success": True, "data": movimientos})

    def _informe_mensual(self, data: dict) -> JsonResponse:
        codigo = data.get("codigo", "").strip()
        ano_raw = data.get("ano", "").strip()
        if not codigo or not ano_raw:
            return JsonResponse({"success": False, "message": "Código de artículo y Año requeridos"})

        try:
            ano = int(ano_raw)
        except ValueError:
            return JsonResponse({"success": False, "message": "Año inválido"})

        detalles = (
            Movs.objects
            .select_related("codigo")
            .filter(codigo=codigo, tipo=7, linea__gt=0, fecha__year=ano)
            .order_by("rut", "fecha")
        )

        proveedores = {}
        proveedores_nombres = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_nombres[p["rut"]] = p["nombre"]

        for d in detalles.iterator():
            if not d.rut:
                continue
            rut = d.rut
            if rut not in proveedores:
                proveedores[rut] = {
                    "rut": rut,
                    "nombre": proveedores_nombres.get(rut, ""),
                    "meses": {m: {"cant": 0, "valor": 0} for m in range(1, 13)},
                }
            mes = d.fecha.month
            cant = float(d.cantidad or 0)
            proveedores[rut]["meses"][mes]["cant"] += cant
            proveedores[rut]["meses"][mes]["valor"] += round(cant * float(d.punit or 0), 0)

        resultado = []
        for rut in sorted(proveedores.keys()):
            prov = proveedores[rut]
            row = {"rut": prov["rut"], "nombre": prov["nombre"]}
            tot_cant = 0
            tot_valor = 0
            for m in range(1, 13):
                cant = round(prov["meses"][m]["cant"], 3)
                valor = int(prov["meses"][m]["valor"])
                row[f"m{m}_cant"] = cant
                row[f"m{m}_valor"] = valor
                tot_cant += cant
                tot_valor += valor
            row["tot_cant"] = round(tot_cant, 3)
            row["tot_valor"] = tot_valor
            resultado.append(row)

        return JsonResponse({"success": True, "data": resultado})

    def _generar_pdf_info(self, request: HttpRequest, data: dict) -> HttpResponse:
        codigo = data.get("codigo", "").strip()
        if not codigo:
            return JsonResponse({"success": False, "message": "Código de artículo requerido"})

        fi, fc = self._parse_fechas(data)
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        art_nombre = ""
        try:
            art = Articulos.objects.get(codigo=codigo)
            art_nombre = art.descr
        except Articulos.DoesNotExist:
            pass

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(codigo=codigo, tipo__isnull=False)
        )
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("rut", "fecha")

        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}
        proveedores_map = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_map[p["rut"]] = p["nombre"]

        grouped = self._agrupar_por_rut(qs, bodegas_map, proveedores_map)

        total_entradas = 0
        total_salidas = 0
        total_monto = 0
        for item in grouped:
            if item["_subtotal"]:
                total_entradas += item["_entradas"]
                total_salidas += item["_salidas"]
                total_monto += item["_monto"]

        saldo_final = total_entradas - total_salidas

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        subtotal_style_pdf = ParagraphStyle(
            "Subtotal", parent=getSampleStyleSheet()["Normal"],
            fontSize=7, leading=9, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#4b5563"),
        )
        subtotal_right = ParagraphStyle("SubR", parent=subtotal_style_pdf, alignment=2)

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=6, leading=9,
            )
            right_style = ParagraphStyle(
                "RightStyle", parent=cell_style, alignment=2,
            )
            center_style = ParagraphStyle(
                "CenterStyle", parent=cell_style, alignment=1,
            )
            resumen_style = ParagraphStyle(
                "Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1,
            )

            elems = []
            elems.append(Paragraph("Info Proveedor por Artículo", title_style))
            elems.append(Paragraph(
                f"<b>Código:</b> {codigo} &nbsp;&nbsp;&nbsp;"
                f"<b>Nombre:</b> {art_nombre}",
                subtitle_style,
            ))
            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 3 * mm))

            header = ["Fecha", "RUT", "Proveedor", "Cantidad", "P.Unit", "Total", "Número", "Bodega"]
            table_data = [header]
            subtotal_rows = []

            for item in grouped:
                if item["_subtotal"]:
                    table_data.append([
                        Paragraph("", subtotal_style_pdf),
                        Paragraph(f"<b>Subtotal</b>", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph(f"<b>{clq(item['_saldo'])}</b>", subtotal_right),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph(f"<b>{cl(item['_monto'])}</b>", subtotal_right),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                    ])
                    subtotal_rows.append(len(table_data) - 1)
                else:
                    table_data.append([
                        Paragraph(str(item["fecha"]), center_style),
                        Paragraph(str(item["rut"]), center_style),
                        Paragraph(str(item["proveedor_nombre"]), cell_style),
                        Paragraph(f"<b>{clq(item['cantidad'])}</b>", right_style),
                        Paragraph(cl(item["punit"]), right_style),
                        Paragraph(cl(item["total"]), right_style),
                        Paragraph(str(item["numero"]), center_style),
                        Paragraph(str(item["bodega_nombre"]), cell_style),
                    ])

            col_widths = [22*mm, 22*mm, 40*mm, 18*mm, 18*mm, 20*mm, 18*mm, 26*mm]

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]
            for sr in subtotal_rows:
                style_cmds.append(("BACKGROUND", (0, sr), (-1, sr), colors.HexColor("#e5e7eb")))
                style_cmds.append(("FONTNAME", (0, sr), (-1, sr), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)
            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[
                Paragraph(f'<b>Total Entradas</b><br/><font color="#16a34a">{clq(total_entradas)}</font>', resumen_style),
                Paragraph(f'<b>Total Salidas</b><br/><font color="#dc2626">{clq(total_salidas)}</font>', resumen_style),
                Paragraph(f'<b>Saldo</b><br/><font color="#2563eb">{clq(saldo_final)}</font>', resumen_style),
                Paragraph(f'<b>Monto Total</b><br/><font color="#7c3aed">{cl(total_monto)}</font>', resumen_style),
            ]]

            resumen_tbl = Table(resumen_data, colWidths=[32*mm, 32*mm, 32*mm, 32*mm], hAlign="CENTER")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elems.append(resumen_tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w/2, 10*mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w-15*mm, 10*mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=22*mm, bottomMargin=18*mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = f"info_proveedor_{codigo}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    def _generar_excel_info(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        codigo = data.get("codigo", "").strip()
        if not codigo:
            return JsonResponse({"success": False, "message": "Código de artículo requerido"})

        fi, fc = self._parse_fechas(data)

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(codigo=codigo, tipo__isnull=False)
        )
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("rut", "fecha")
        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}
        proveedores_map = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_map[p["rut"]] = p["nombre"]

        grouped = self._agrupar_por_rut(qs, bodegas_map, proveedores_map)

        wb = Workbook()
        ws = wb.active
        ws.title = "Info Proveedor"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Fecha", "RUT", "Proveedor", "Cantidad", "P.Unit", "Total", "Tipo", "Número", "Bodega"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        sub_font = Font(name="Calibri", bold=True, size=10)
        sub_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        total_entradas = 0
        total_salidas = 0
        total_monto = 0

        for item in grouped:
            if item["_subtotal"]:
                total_entradas += item["_entradas"]
                total_salidas += item["_salidas"]
                total_monto += item["_monto"]

                sub_row = ["", f"Subtotal {item['_rut']}", "", item["_saldo"], "", item["_monto"], "", "", ""]
                ws.append(sub_row)
                row_num = ws.max_row
                for col_idx in range(1, len(sub_row) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = sub_font
                    cell.fill = sub_fill
                    cell.border = thin_border
                    if col_idx in (4, 5, 6):
                        cell.alignment = num_align
                        cell.number_format = '#,##0.000' if col_idx == 4 else '#,##0'
                    else:
                        cell.alignment = center_align
            else:
                row_data = [
                    item["fecha"],
                    item["rut"],
                    item["proveedor_nombre"],
                    item["cantidad"],
                    item["punit"],
                    item["total"],
                    item["tipo_nombre"],
                    item["numero"],
                    item["bodega_nombre"],
                ]
                ws.append(row_data)
                row_num = ws.max_row
                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in (4, 5, 6):
                        cell.alignment = num_align
                        if col_idx == 4:
                            cell.number_format = '#,##0.000'
                        else:
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = center_align

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 22
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 22

        ws.append([])
        saldo_final = total_entradas - total_salidas

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        ws.append([f"Total Entradas: {clq(total_entradas)}  |  Total Salidas: {clq(total_salidas)}  |  Saldo: {clq(saldo_final)}  |  Monto Total: {cl(total_monto)}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"info_proveedor_{codigo}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf_mensual(self, request: HttpRequest, data: dict) -> HttpResponse:
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        codigo = data.get("codigo", "").strip()
        ano_raw = data.get("ano", "").strip()
        if not codigo or not ano_raw:
            return JsonResponse({"success": False, "message": "Código de artículo y Año requeridos"})

        ano = int(ano_raw)

        art_nombre = ""
        try:
            art = Articulos.objects.get(codigo=codigo)
            art_nombre = art.descr
        except Articulos.DoesNotExist:
            pass

        proveedores_nombres = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_nombres[p["rut"]] = p["nombre"]

        nombres_meses_abr = [
            "Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"
        ]

        detalles = (
            Movs.objects
            .select_related("codigo")
            .filter(codigo=codigo, tipo=7, linea__gt=0, fecha__year=ano)
            .order_by("rut", "fecha")
        )

        proveedores = {}
        for d in detalles.iterator():
            if not d.rut:
                continue
            rut = d.rut
            if rut not in proveedores:
                proveedores[rut] = {
                    "rut": rut,
                    "nombre": proveedores_nombres.get(rut, ""),
                    "meses": {m: {"cant": 0, "valor": 0} for m in range(1, 13)},
                }
            mes = d.fecha.month
            cant = float(d.cantidad or 0)
            proveedores[rut]["meses"][mes]["cant"] += cant
            proveedores[rut]["meses"][mes]["valor"] += round(cant * float(d.punit or 0), 0)

        data_rows = []
        for rut in sorted(proveedores.keys()):
            prov = proveedores[rut]
            cant_row = [prov["rut"], prov["nombre"]]
            val_row = [prov["rut"], prov["nombre"]]
            tot_cant = 0
            tot_valor = 0
            for m in range(1, 13):
                cant = round(prov["meses"][m]["cant"], 3)
                valor = int(prov["meses"][m]["valor"])
                cant_row.append(cant)
                val_row.append(valor)
                tot_cant += cant
                tot_valor += valor
            cant_row.append(round(tot_cant, 3))
            val_row.append(tot_valor)
            data_rows.append((cant_row, val_row))

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_table(title_text, is_valor):
            styles = getSampleStyleSheet()
            cell_style = ParagraphStyle("Cell", fontSize=7, leading=9)
            right_style = ParagraphStyle("Right", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("Center", parent=cell_style, alignment=1)
            bold_style = ParagraphStyle("Bold", parent=cell_style, fontName="Helvetica-Bold")
            bold_right = ParagraphStyle("BoldRight", parent=bold_style, alignment=2)
            header_style = ParagraphStyle("Hdr", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)

            hdr = [Paragraph("<b>RUT</b>", header_style), Paragraph("<b>Proveedor</b>", header_style)]
            for mn in nombres_meses_abr:
                hdr.append(Paragraph(f"<b>{mn}</b>", header_style))
            hdr.append(Paragraph("<b>Total</b>", header_style))

            table_data = [hdr]
            for cr, vr in data_rows:
                row = [Paragraph(str(cr[0]), center_style), Paragraph(str(cr[1]), cell_style)]
                src = vr if is_valor else cr
                for i in range(2, len(src)):
                    if not is_valor:
                        row.append(Paragraph(clq(src[i]), right_style))
                    else:
                        row.append(Paragraph(f"${cl(src[i])}", right_style))
                table_data.append(row)

            tot_row = [Paragraph("<b>TOTAL</b>", bold_style), Paragraph("", cell_style)]
            monthly_sums = []
            for m in range(1, 13):
                s = sum((vr if is_valor else cr)[m + 1] for cr, vr in data_rows)
                if not is_valor:
                    s = round(s, 3)
                    tot_row.append(Paragraph(f"<b>{clq(s)}</b>", bold_right))
                else:
                    s = int(s)
                    tot_row.append(Paragraph(f"<b>${cl(s)}</b>", bold_right))
                monthly_sums.append(s)
            grand = sum(monthly_sums)
            if not is_valor:
                tot_row.append(Paragraph(f"<b>{clq(grand)}</b>", bold_right))
            else:
                tot_row.append(Paragraph(f"<b>${cl(grand)}</b>", bold_right))
            table_data.append(tot_row)

            col_widths = [16*mm, 26*mm] + [12*mm] * 12 + [14*mm]
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e5e7eb")),
            ]))
            return tbl

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w/2, 10*mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w-15*mm, 10*mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        buf = io.BytesIO()

        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=10*mm, rightMargin=10*mm,
            topMargin=18*mm, bottomMargin=18*mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4*mm, fontSize=13)
        subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3*mm, fontSize=9)

        elems = [
            Paragraph("Informe Mensual por Artículo", title_style),
            Paragraph(f"<b>Código:</b> {codigo} &nbsp;&nbsp;&nbsp; <b>Nombre:</b> {art_nombre} &nbsp;&nbsp;&nbsp; <b>Año:</b> {ano}", subtitle_style),
            Spacer(1, 3*mm),
            Paragraph("<b>Cantidades</b>", subtitle_style),
            build_table("Cantidades", False),
            Spacer(1, 5*mm),
            Paragraph("<b>Valores ($)</b>", subtitle_style),
            build_table("Valores", True),
            Spacer(1, 3*mm),
            Paragraph(f"<i>Generado por {usuario} el {now_str}</i>", ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.gray)),
        ]
        doc.build(elems, canvasmaker=_Canvas)
        buf.seek(0)

        filename = f"informe_mensual_{codigo}_{ano}.pdf"
        response = HttpResponse(buf.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_excel_mensual(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        codigo = data.get("codigo", "").strip()
        ano_raw = data.get("ano", "").strip()
        if not codigo or not ano_raw:
            return JsonResponse({"success": False, "message": "Código de artículo y Año requeridos"})

        ano = int(ano_raw)

        proveedores_nombres = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_nombres[p["rut"]] = p["nombre"]

        detalles = (
            Movs.objects
            .select_related("codigo")
            .filter(codigo=codigo, tipo=7, linea__gt=0, fecha__year=ano)
            .order_by("rut", "fecha")
        )

        proveedores = {}
        for d in detalles.iterator():
            if not d.rut:
                continue
            rut = d.rut
            if rut not in proveedores:
                proveedores[rut] = {
                    "rut": rut,
                    "nombre": proveedores_nombres.get(rut, ""),
                    "meses": {m: {"cant": 0, "valor": 0} for m in range(1, 13)},
                }
            mes = d.fecha.month
            cant = float(d.cantidad or 0)
            proveedores[rut]["meses"][mes]["cant"] += cant
            proveedores[rut]["meses"][mes]["valor"] += round(cant * float(d.punit or 0), 0)

        wb = Workbook()

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")

        month_names = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

        for sheet_name, is_valor in [("Cantidades", False), ("Valores", True)]:
            ws = wb.create_sheet(title=sheet_name) if sheet_name == "Valores" else wb.active
            ws.title = sheet_name

            hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="d1d5db"),
                right=Side(style="thin", color="d1d5db"),
                top=Side(style="thin", color="d1d5db"),
                bottom=Side(style="thin", color="d1d5db"),
            )

            headers = ["RUT", "Proveedor"] + month_names + ["Total"]
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = thin_border

            data_font = Font(name="Calibri", size=10)
            num_align = Alignment(horizontal="right", vertical="center")
            center_align = Alignment(horizontal="center", vertical="center")

            grand_tot = 0
            monthly_grands = [0] * 12

            for rut in sorted(proveedores.keys()):
                prov = proveedores[rut]
                row = [prov["rut"], prov["nombre"]]
                row_tot = 0
                for m in range(1, 13):
                    val = prov["meses"][m]["valor"] if is_valor else round(prov["meses"][m]["cant"], 3)
                    if is_valor:
                        val = int(val)
                        row.append(val)
                    else:
                        row.append(val)
                    row_tot += val
                    monthly_grands[m - 1] += val
                if is_valor:
                    row.append(row_tot)
                else:
                    row.append(round(row_tot, 3))
                grand_tot += row_tot
                ws.append(row)
                row_num = ws.max_row
                for col_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx >= 3:
                        cell.alignment = num_align
                        if is_valor and col_idx > 2:
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = center_align

            tot_row = ["TOTAL", ""] + monthly_grands + [grand_tot]
            ws.append(tot_row)
            row_num = ws.max_row
            sub_font = Font(name="Calibri", bold=True, size=10)
            sub_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
            for col_idx in range(1, len(tot_row) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = sub_font
                cell.fill = sub_fill
                cell.border = thin_border
                if col_idx >= 3:
                    cell.alignment = num_align
                    if is_valor and col_idx > 2:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = center_align

            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 30
            for i in range(3, 16):
                ws.column_dimensions[chr(64 + i)].width = 12

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"informe_mensual_{codigo}_{ano}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
