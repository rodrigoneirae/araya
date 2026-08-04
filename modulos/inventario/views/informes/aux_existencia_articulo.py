import io
import os
from typing import Any
from datetime import datetime, timedelta
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponseRedirect, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.articulos import Articulos
from modulos.maestros.models.bodegas import Bodegas


class IndexInformeAuxExistenciaArticuloView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/inventario/informes/aux_existencia_articulo.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        if action == "buscar_articulo":
            return self._buscar_articulo(request.POST.get("codigo"))
        elif action == "listar_articulos":
            return self._listar_articulos()
        elif action == "listar_bodegas":
            return self._listar_bodegas()
        elif action == "consultar":
            return self._consultar(request.POST)
        elif action == "generar_pdf":
            return self._generar_pdf(request)
        elif action == "generar_excel":
            return self._generar_excel(request)
        return JsonResponse({"success": False})

    def _buscar_articulo(self, codigo: str | None) -> JsonResponse:
        if not codigo:
            return JsonResponse({"success": False})
        try:
            art = Articulos.objects.get(codigo=codigo.strip())
            return JsonResponse({
                "success": True,
                "data": {
                    "codigo": art.codigo,
                    "nombre": art.descr or "",
                    "um": art.um or "",
                }
            })
        except Articulos.DoesNotExist:
            return JsonResponse({"success": False, "message": "Artículo no encontrado"})

    def _listar_articulos(self) -> JsonResponse:
        articulos = Articulos.objects.values("codigo", "descr", "um").order_by("descr")
        return JsonResponse({"articulos": list(articulos)})

    def _listar_bodegas(self) -> JsonResponse:
        bodegas = Bodegas.objects.values("cod", "nombre").order_by("nombre")
        return JsonResponse({"bodegas": list(bodegas)})

    def _consultar(self, data) -> JsonResponse:

        codigo = data.get("codigo", "").strip()
        fecha_desde = data.get("fecha_desde", "").strip()


        if not codigo:
            return JsonResponse({
                "success": False,
                "message": "Debe ingresar un código de artículo"
            })

        fc = None

        if fecha_desde:
            try:
                fc = datetime.strptime(
                    fecha_desde.strip(),
                    "%Y-%m-%d"
                )

                print("fc:", fc)

            except Exception as e:
                print("ERROR FECHA:", e)

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(
                codigo__codigo=codigo,
                tipo__isnull=False,
                tipo__signo__isnull=False,
                linea__isnull=False,
            )
            .exclude(tipo__signo=0)
            .exclude(linea=0)
        )

        # IMPORTANTE: aplicar fecha aquí
        if fc:
            print(fc)
            qs = qs.filter(fecha__lt=fc)

        qs = qs.distinct().order_by("fecha")

        print("Registros:", qs.count())
        print(qs.query)

        bodegas_map = {
            b['cod']: b['nombre']
            for b in Bodegas.objects.values('cod', 'nombre')
        }

        resultados = []

        for m in qs.iterator():
            resultados.append({
                "codigo": m.codigo.codigo if m.codigo else "",
                "descr": m.codigo.descr if m.codigo else "",
                "fecha": m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                "numero": m.numero,
                "bodega": bodegas_map.get(
                    m.bodega,
                    f"Bodega {m.bodega}" if m.bodega else ""
                ),
                "cantidad": m.cantidad or 0,
                "tipo": m.tipo.nombre if m.tipo else "",
                "signo": m.tipo.signo if m.tipo else 0,
            })

        return JsonResponse({
            "success": True,
            "data": resultados,
        })

    def _generar_excel(self, request: HttpRequest) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        data = request.POST
        codigo = data.get("codigo", "").strip()
        fecha_desde = data.get("fecha_desde", "").strip()
        fecha_corte = data.get("fecha_corte", "").strip()
        saldo_ant = data.get("saldo_ant", "0").strip()

        try:
            saldo_ant = float(saldo_ant)
        except Exception:
            saldo_ant = 0

        if not codigo:
            return JsonResponse({"success": False, "message": "Debe ingresar un código de artículo"})

        fd = fc = None
        if fecha_desde:
            try: fd = datetime.strptime(fecha_desde.strip(), "%Y-%m-%d")
            except: pass
        if fecha_corte:
            try: fc = datetime.strptime(fecha_corte.strip(), "%Y-%m-%d")
            except: pass

        qs = (
            Movs.objects.select_related("tipo", "codigo")
            .filter(codigo__codigo=codigo, tipo__isnull=False, tipo__signo__isnull=False, linea__isnull=False)
            .exclude(tipo__signo=0).exclude(linea=0)
        )
        if fd: qs = qs.filter(fecha__gte=fd)
        if fc: qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))
        qs = qs.distinct().order_by("fecha")

        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}

        rows = []
        total_entradas = total_salidas = 0
        for m in qs.iterator():
            signo = m.tipo.signo if m.tipo else 0
            cantidad = m.cantidad or 0
            if signo > 0: total_entradas += cantidad
            if signo < 0: total_salidas += abs(cantidad)
            rows.append([
                m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                cantidad,
                bodegas_map.get(m.bodega, str(m.bodega or "")),
                str(int(m.numero)) if m.numero else "",
                m.tipo.nombre if m.tipo else "",
                m.punit or 0,
            ])

        saldo_final = saldo_ant
        total_linea = total_entradas + saldo_final

        wb = Workbook()
        ws = wb.active
        ws.title = "Auxiliar Existencia"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Fecha", "Cantidad", "Bodega", "Número", "Documento", "P.Unitario"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        for r in rows:
            ws.append(r)
            row_num = ws.max_row
            for col_idx in range(1, len(r) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in (2, 6):
                    cell.alignment = num_align
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = center_align

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 32
        ws.column_dimensions["F"].width = 14

        ws.append([])
        res_font = Font(name="Calibri", bold=True, size=10)
        res_align = Alignment(horizontal="center", vertical="center")
        green_font = Font(name="Calibri", bold=True, size=10, color="16a34a")
        blue_font = Font(name="Calibri", bold=True, size=10, color="2563eb")
        dark_font = Font(name="Calibri", bold=True, size=10, color="1f2937")

        def cl(v):
            return f"{float(v):,.0f}".replace(",", ".")

        ws.append([f"Total Entradas: {cl(total_entradas)}   |   Saldo Final: {cl(saldo_final)}   |   Total: {cl(total_linea)}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = res_font
        cell.alignment = res_align

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"aux_existencia_{codigo}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf(self, request: HttpRequest) -> HttpResponse:

        data = request.POST

        codigo = data.get("codigo", "").strip()
        fecha_desde = data.get("fecha_desde", "").strip()
        fecha_corte = data.get("fecha_corte", "").strip()

        saldo_ant = data.get("saldo_ant", "0").strip()

        try:
            saldo_ant = float(saldo_ant)
        except Exception:
            saldo_ant = 0

        usuario = str(request.user)

        if not codigo:
            return JsonResponse({
                "success": False,
                "message": "Debe ingresar un código de artículo"
            })

        fd = None
        fc = None

        if fecha_desde:
            try:
                fd = datetime.strptime(
                    fecha_desde.strip(),
                    "%Y-%m-%d"
                )
            except Exception:
                pass

        if fecha_corte:
            try:
                fc = datetime.strptime(
                    fecha_corte.strip(),
                    "%Y-%m-%d"
                )
            except Exception:
                pass

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(
                codigo__codigo=codigo,
                tipo__isnull=False,
                tipo__signo__isnull=False,
                linea__isnull=False,
            )
            .exclude(tipo__signo=0)
            .exclude(linea=0)
        )

        if fd:
            qs = qs.filter(fecha__gte=fd)

        if fc:
            qs = qs.filter(
                fecha__lte=fc.replace(
                    hour=23,
                    minute=59,
                    second=59
                )
            )

        qs = qs.distinct().order_by("fecha")

        bodegas_map = {
            b["cod"]: b["nombre"]
            for b in Bodegas.objects.values("cod", "nombre")
        }

        rows = []

        total_entradas = 0
        total_salidas = 0

        for m in qs.iterator():

            signo = m.tipo.signo if m.tipo else 0

            cantidad = m.cantidad or 0

            if signo > 0:
                total_entradas += cantidad

            if signo < 0:
                total_salidas += abs(cantidad)

            rows.append([
                m.fecha.strftime("%d-%m-%Y")
                if m.fecha else "",

                cantidad,

                bodegas_map.get(
                    m.bodega,
                    str(m.bodega or "")
                ),

                str(int(m.numero))
                if m.numero else "",

                m.tipo.nombre
                if m.tipo else "",

                m.punit or 0,
            ])

        saldo_final = saldo_ant
        total_linea = saldo_ant + total_entradas - total_salidas

        try:
            art_info = Articulos.objects.get(codigo=codigo)

            articulo_nombre = art_info.descr or ""
            um = art_info.um or ""

        except Articulos.DoesNotExist:

            articulo_nombre = ""
            um = ""

        fecha_inf = ""

        if fd and fc:
            fecha_inf = (
                f"desde {fd.strftime('%d-%m-%Y')} "
                f"hasta {fc.strftime('%d-%m-%Y')}"
            )

        elif fd:
            fecha_inf = f"desde {fd.strftime('%d-%m-%Y')}"

        elif fc:
            fecha_inf = f"hasta {fc.strftime('%d-%m-%Y')}"

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )

        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{float(v):,.0f}".replace(",", ".")

        def build_elements():

            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading2"],
                spaceAfter=4 * mm,
                fontSize=14
            )

            subtitle_style = ParagraphStyle(
                "CustomSub",
                parent=styles["Normal"],
                spaceAfter=3 * mm,
                fontSize=9,
            )

            cell_style = ParagraphStyle(
                "CellStyle",
                parent=styles["Normal"],
                fontSize=7,
                leading=9,
            )

            # DERECHA
            right_style = ParagraphStyle(
                "RightStyle",
                parent=cell_style,
                alignment=2,
            )

            # CENTRO
            center_style = ParagraphStyle(
                "CenterStyle",
                parent=cell_style,
                alignment=1,
            )

            resumen_style = ParagraphStyle(
                "Resumen",
                parent=styles["Normal"],
                fontSize=8,
                leading=11,
                alignment=1,
            )

            elems = []

            elems.append(
                Paragraph(
                    "Auxiliar de Existencia por Artículo",
                    title_style
                )
            )

            elems.append(
                Paragraph(
                    f"<b>Código:</b> {codigo} "
                    f"&nbsp;&nbsp;&nbsp;"
                    f"<b>Nombre:</b> {articulo_nombre} "
                    f"&nbsp;&nbsp;&nbsp;"
                    f"<b>UM:</b> {um}",
                    subtitle_style,
                )
            )

            if fecha_inf:
                elems.append(
                    Paragraph(
                        f"<b>Período:</b> {fecha_inf}",
                        subtitle_style
                    )
                )

            elems.append(Spacer(1, 3 * mm))

            header = [
                "Fecha",
                "Cantidad",
                "Bodega",
                "Número",
                "Documento",
                "P.Unitario"
            ]

            table_data = [header]

            for r in rows:
                table_data.append([

                    # FECHA
                    Paragraph(
                        str(r[0]),
                        center_style
                    ),

                    # CANTIDAD DERECHA
                    Paragraph(
                        f"<b>{cl(r[1])}</b>",
                        right_style
                    ),

                    # BODEGA
                    Paragraph(
                        str(r[2]),
                        cell_style
                    ),

                    # NUMERO CENTRADO
                    Paragraph(
                        str(r[3]),
                        center_style
                    ),

                    # DOCUMENTO CENTRADO
                    Paragraph(
                        str(r[4]),
                        center_style
                    ),

                    # P.UNITARIO DERECHA
                    Paragraph(
                        cl(r[5]),
                        right_style
                    ),
                ])

            col_widths = [
                28 * mm,
                22 * mm,
                45 * mm,
                25 * mm,
                45 * mm,
                25 * mm
            ]

            tbl = Table(
                table_data,
                colWidths=col_widths,
                repeatRows=1
            )

            tbl.setStyle(TableStyle([

                # HEADER
                ("BACKGROUND", (0, 0), (-1, 0),
                 colors.HexColor("#1f2937")),

                ("TEXTCOLOR", (0, 0), (-1, 0),
                 colors.white),

                ("FONTNAME", (0, 0), (-1, 0),
                 "Helvetica-Bold"),

                # GENERAL
                ("VALIGN", (0, 0), (-1, -1),
                 "MIDDLE"),

                ("FONTSIZE", (0, 0), (-1, -1),
                 7),

                ("GRID", (0, 0), (-1, -1),
                 0.5, colors.HexColor("#d1d5db")),

                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white,
                  colors.HexColor("#f9fafb")]),

                ("TOPPADDING", (0, 0), (-1, -1),
                 4),

                ("BOTTOMPADDING", (0, 0), (-1, -1),
                 4),

                # HEADER CENTRADO
                ("ALIGN", (0, 0), (-1, 0),
                 "CENTER"),

            ]))

            elems.append(tbl)

            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[

                Paragraph(
                    f'<b>Total Entradas</b><br/>'
                    f'<font color="#16a34a">{cl(total_entradas)}</font>',
                    resumen_style
                ),

                Paragraph(
                    f'<b>Saldo Anterior</b><br/>'
                    f'<font color="#2563eb">{cl(saldo_final)}</font>',
                    resumen_style
                ),

                Paragraph(
                    f'<b>Total Final</b><br/>'
                    f'<font color="#111827">{cl(total_linea)}</font>',
                    resumen_style
                ),
            ]]

            resumen_tbl = Table(
                resumen_data,
                colWidths=[
                    42 * mm,
                    42 * mm,
                    42 * mm
                ],
                hAlign="CENTER"
            )

            resumen_tbl.setStyle(TableStyle([

                ("BACKGROUND", (0, 0), (-1, -1),
                 colors.HexColor("#f3f4f6")),

                ("BOX", (0, 0), (-1, -1),
                 0.8, colors.HexColor("#d1d5db")),

                ("INNERGRID", (0, 0), (-1, -1),
                 0.5, colors.HexColor("#d1d5db")),

                ("ALIGN", (0, 0), (-1, -1),
                 "CENTER"),

                ("VALIGN", (0, 0), (-1, -1),
                 "MIDDLE"),

                ("TOPPADDING", (0, 0), (-1, -1),
                 8),

                ("BOTTOMPADDING", (0, 0), (-1, -1),
                 8),

            ]))

            elems.append(resumen_tbl)

            return elems

        class _Canvas(rl_canvas.Canvas):

            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):

                total = len(self._saved)

                for state in self._saved:

                    self.__dict__.update(state)

                    w, h = self._pagesize

                    if logo_path:

                        try:

                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader

                            _img = PILImage.open(logo_path)

                            if _img.mode == "RGBA":

                                _bg = PILImage.new(
                                    "RGB",
                                    _img.size,
                                    (255, 255, 255)
                                )

                                _bg.paste(
                                    _img,
                                    mask=_img.split()[3]
                                )

                                self.drawImage(
                                    ImageReader(_bg),
                                    15 * mm,
                                    h - 17 * mm,
                                    width=50 * mm,
                                    height=11 * mm,
                                    preserveAspectRatio=True
                                )

                            else:

                                self.drawImage(
                                    ImageReader(_img),
                                    15 * mm,
                                    h - 17 * mm,
                                    width=50 * mm,
                                    height=11 * mm,
                                    preserveAspectRatio=True
                                )

                        except Exception:
                            pass

                    self.setFont("Helvetica", 7)

                    self.setFillColor(
                        colors.HexColor("#6b7280")
                    )

                    self.drawCentredString(
                        w / 2,
                        10 * mm,
                        f"Página {self.getPageNumber()} de {total}"
                    )

                    self.drawRightString(
                        w - 15 * mm,
                        10 * mm,
                        f"Usuario: {usuario} - {now_str}"
                    )

                    super().showPage()

                super().save()

        buf = io.BytesIO()

        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=22 * mm,
            bottomMargin=18 * mm,
        )

        doc.build(
            build_elements(),
            canvasmaker=_Canvas
        )

        pdf_bytes = buf.getvalue()

        buf.close()

        filename = f"aux_existencia_{codigo}.pdf"

        response = HttpResponse(
            pdf_bytes,
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            f'inline; filename="{filename}"'
        )

        return response
