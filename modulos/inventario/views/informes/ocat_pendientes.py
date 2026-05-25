import io
import os
from typing import Any
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.prov_cliente import Provclientes


class IndexInformeOcatPendientesView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/inventario/informes/ocat_pendientes.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        handlers = {
            "listar_proveedores": lambda _: self._listar_proveedores(),
            "buscar_proveedor": lambda d: self._buscar_proveedor(d.get("rut")),
            "info_informe": lambda d: self._info_informe(d),
            "info_resumen": lambda d: self._info_resumen(d),
            "generar_pdf": lambda d: self._generar_pdf(request, d),
            "generar_excel": lambda d: self._generar_excel(request, d),
            "generar_pdf_resumen": lambda d: self._generar_pdf_resumen(request, d),
            "generar_excel_resumen": lambda d: self._generar_excel_resumen(request, d),
        }
        handler = handlers.get(action, lambda _: JsonResponse({"success": False}))
        return handler(request.POST)

    def _parse_fechas(self, data: dict) -> tuple:
        fi = None
        fc = None
        fecha_inicio = data.get("fecha_inicio", "").strip()
        fecha_corte = data.get("fecha_corte", "").strip()
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            except Exception:
                pass
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte, "%Y-%m-%d")
            except Exception:
                pass
        return fi, fc

    def _listar_proveedores(self) -> JsonResponse:
        proveedores = (
            Provclientes.objects
            .exclude(rut='')
            .values("rut", "nombre")
            .order_by("nombre")
        )
        return JsonResponse({"proveedores": list(proveedores)})

    def _buscar_proveedor(self, rut: str | None) -> JsonResponse:
        if not rut:
            return JsonResponse({"success": False})
        try:
            prov = Provclientes.objects.get(rut=rut.strip())
            return JsonResponse({
                "success": True,
                "data": {"rut": prov.rut, "nombre": prov.nombre},
            })
        except Provclientes.DoesNotExist:
            return JsonResponse({"success": False, "message": "Proveedor no encontrado"})

    def _get_proveedores_map(self) -> dict:
        return {p["rut"]: p["nombre"] for p in Provclientes.objects.values("rut", "nombre")}

    def _build_result(self, fi, fc, rut: str = "") -> list[dict]:
        headers_qs = Movs.objects.filter(tipo=7, linea=0)
        details_qs = Movs.objects.filter(tipo=7, linea=1).select_related("codigo")

        if rut:
            headers_qs = headers_qs.filter(rut=rut)
            details_qs = details_qs.filter(rut=rut)
        if fi:
            headers_qs = headers_qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
            details_qs = details_qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            headers_qs = headers_qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))
            details_qs = details_qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        headers_qs = headers_qs.order_by("numero")
        details_qs = details_qs.order_by("numero")

        proveedores_map = self._get_proveedores_map()

        details_by_numero = {}
        for d in details_qs.iterator():
            num = int(d.numero) if d.numero else 0
            if num not in details_by_numero:
                details_by_numero[num] = []
            details_by_numero[num].append({
                "codigo": d.codigo.codigo if d.codigo else "",
                "articulo": d.codigo.descr if d.codigo else "",
                "cantidad": d.cantidad or 0,
                "punit": d.punit or 0,
                "total": (d.cantidad or 0) * (d.punit or 0),
            })

        movimientos = []
        for h in headers_qs.iterator():
            num = int(h.numero) if h.numero else 0
            detalles = details_by_numero.get(num, [])

            sub_cant = 0
            sub_total = 0
            first = True
            for det in detalles:
                cant = det["cantidad"]
                total = det["total"]
                sub_cant += cant
                sub_total += total

                movimientos.append({
                    "_subtotal": False,
                    "_first_in_group": first,
                    "numero": str(num),
                    "fecha": h.fecha.strftime("%d-%m-%Y") if h.fecha else "",
                    "rut": h.rut or "",
                    "proveedor": proveedores_map.get(h.rut, ""),
                    "canttotal": h.canttotal or 0,
                    "estado": h.estado or "",
                    "codigo": det["codigo"],
                    "articulo": det["articulo"],
                    "cantidad": cant,
                    "punit": det["punit"],
                    "total": total,
                })
                first = False

            if not detalles:
                movimientos.append({
                    "_subtotal": False,
                    "_first_in_group": True,
                    "numero": str(num),
                    "fecha": h.fecha.strftime("%d-%m-%Y") if h.fecha else "",
                    "rut": h.rut or "",
                    "proveedor": proveedores_map.get(h.rut, ""),
                    "canttotal": h.canttotal or 0,
                    "estado": h.estado or "",
                    "codigo": "",
                    "articulo": "(sin detalle)",
                    "cantidad": 0,
                    "punit": 0,
                    "total": 0,
                })

            movimientos.append({
                "_subtotal": True,
                "_numero": num,
                "_fecha": h.fecha.strftime("%d-%m-%Y") if h.fecha else "",
                "_proveedor": proveedores_map.get(h.rut, ""),
                "_cantidad": sub_cant,
                "_total": sub_total,
            })

        return movimientos

    def _info_informe(self, data: dict) -> JsonResponse:
        fi, fc = self._parse_fechas(data)
        rut = data.get("rut", "").strip()
        movimientos = self._build_result(fi, fc, rut)
        return JsonResponse({"success": True, "data": movimientos})

    def _build_resumen(self, fi, fc, rut: str = "") -> list[dict]:
        headers_qs = Movs.objects.filter(tipo=7, linea=0)
        if rut:
            headers_qs = headers_qs.filter(rut=rut)
        if fi:
            headers_qs = headers_qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            headers_qs = headers_qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))
        headers_qs = headers_qs.order_by("numero")

        nums = [int(h.numero) for h in headers_qs.iterator() if h.numero]
        if not nums:
            return []

        from django.db.models import Sum
        sums = (
            Movs.objects
            .filter(tipo=7, linea=1, numero__in=nums)
            .values("numero")
            .annotate(cant_sum=Sum("cantidad"))
        )
        sum_map = {}
        for s in sums:
            sum_map[int(s["numero"])] = float(s["cant_sum"] or 0)

        proveedores_map = self._get_proveedores_map()
        resumen = []
        for h in headers_qs.iterator():
            num = int(h.numero) if h.numero else 0
            cant_total = h.canttotal or 0
            cant_sum = sum_map.get(num, 0)
            faltante = cant_sum - cant_total
            resumen.append({
                "numero": str(num),
                "fecha": h.fecha.strftime("%d-%m-%Y") if h.fecha else "",
                "rut": h.rut or "",
                "proveedor": proveedores_map.get(h.rut, ""),
                "canttotal": cant_total,
                "cant_sum": cant_sum,
                "faltante": faltante,
            })
        return resumen

    def _info_resumen(self, data: dict) -> JsonResponse:
        fi, fc = self._parse_fechas(data)
        rut = data.get("rut", "").strip()
        resumen = self._build_resumen(fi, fc, rut)
        return JsonResponse({"success": True, "data": resumen})

    @staticmethod
    def _cl(v):
        return f"{int(round(float(v))):,}".replace(",", ".")

    @staticmethod
    def _clq(v):
        val = float(v)
        return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

    # --- PDF generation ---
    def _generar_pdf(self, request: HttpRequest, data: dict) -> HttpResponse:
        fi, fc = self._parse_fechas(data)
        rut = data.get("rut", "").strip()
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        proveedores_map = self._get_proveedores_map()
        movimientos = self._build_result(fi, fc, rut)

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        prov_nombre = ""
        if rut:
            try:
                prov = Provclientes.objects.get(rut=rut)
                prov_nombre = prov.nombre
            except Provclientes.DoesNotExist:
                pass

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=2 * mm, fontSize=8,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=5.5, leading=8,
            )
            right_style = ParagraphStyle(
                "RightStyle", parent=cell_style, alignment=2,
            )
            center_style = ParagraphStyle(
                "CenterStyle", parent=cell_style, alignment=1,
            )
            resumen_style = ParagraphStyle(
                "Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1,
            )
            sub_style = ParagraphStyle(
                "Subtotal", parent=getSampleStyleSheet()["Normal"],
                fontSize=6, leading=9, fontName="Helvetica-Bold",
                textColor=colors.HexColor("#4b5563"),
            )
            sub_right = ParagraphStyle("SubR", parent=sub_style, alignment=2)

            elems = []
            elems.append(Paragraph("OC/AT Pendientes", title_style))
            if rut:
                elems.append(Paragraph(
                    f"<b>RUT:</b> {rut} &nbsp;&nbsp;&nbsp;"
                    f"<b>Nombre:</b> {prov_nombre}",
                    subtitle_style,
                ))
            else:
                elems.append(Paragraph("<b>Todos los proveedores</b>", subtitle_style))

            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 2 * mm))

            headers = ["N° OC", "Fecha", "RUT", "Proveedor", "Artículo", "Cant.", "P.Unit", "Total"]
            col_widths = [18 * mm, 18 * mm, 20 * mm, 38 * mm, 42 * mm, 16 * mm, 16 * mm, 18 * mm]
            table_data = [headers]
            subtotal_rows = []
            total_cant_pdf = 0
            total_monto_pdf = 0

            sub_idx = 0
            sub_info = []
            rows_buffer = []

            for m in movimientos:
                if m.get("_subtotal"):
                    sub_info.append(m)
                else:
                    rows_buffer.append(m)

            sub_totals = []
            for s in sub_info:
                sub_totals.append((s["_numero"], s["_cantidad"], s["_total"]))

            total_cant_pdf = sum(st[1] for st in sub_totals)
            total_monto_pdf = sum(st[2] for st in sub_totals)

            st_idx = 0
            for m in movimientos:
                if m.get("_subtotal"):
                    _num = m["_numero"]
                    _cant = m["_cantidad"]
                    _total = m["_total"]
                    table_data.append([
                        Paragraph(f"<b>Subtotal {_num}</b>", sub_style),
                        Paragraph("", sub_style),
                        Paragraph("", sub_style),
                        Paragraph(f"<b>{m['_proveedor']}</b>", sub_style),
                        Paragraph("", sub_style),
                        Paragraph(f"<b>{self._clq(_cant)}</b>", sub_right),
                        Paragraph("", sub_style),
                        Paragraph(f"<b>{self._cl(_total)}</b>", sub_right),
                    ])
                    subtotal_rows.append(len(table_data) - 1)
                    st_idx += 1
                else:
                    art_label = f"{m['codigo']} {m['articulo']}".strip()
                    table_data.append([
                        Paragraph(m["numero"], center_style),
                        Paragraph(m["fecha"], center_style),
                        Paragraph(m["rut"], center_style),
                        Paragraph(m["proveedor"], cell_style),
                        Paragraph(art_label, cell_style),
                        Paragraph(self._clq(m["cantidad"]), right_style),
                        Paragraph(self._cl(m["punit"]), right_style),
                        Paragraph(self._cl(m["total"]), right_style),
                    ])

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]
            for sr in subtotal_rows:
                style_cmds.append(("BACKGROUND", (0, sr), (-1, sr), colors.HexColor("#e5e7eb")))
                style_cmds.append(("FONTNAME", (0, sr), (-1, sr), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)

            elems.append(Spacer(1, 5 * mm))
            resumen_data = [[
                Paragraph(f'<b>Total Cantidad</b><br/><font color="#2563eb">{self._clq(total_cant_pdf)}</font>', resumen_style),
                Paragraph(f'<b>Monto Total</b><br/><font color="#7c3aed">{self._cl(total_monto_pdf)}</font>', resumen_style),
                Paragraph(f'<b>Total Órdenes</b><br/><font color="#4b5563">{len(sub_info)}</font>', resumen_style),
            ]]
            resumen_tbl = Table(resumen_data, colWidths=[42 * mm, 42 * mm, 42 * mm], hAlign="CENTER")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elems.append(resumen_tbl)

            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=12 * mm, rightMargin=12 * mm,
            topMargin=18 * mm, bottomMargin=16 * mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()
        filename = f"ocat_pendientes_{rut or 'todos'}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    # --- Excel generation ---
    def _generar_excel(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        fi, fc = self._parse_fechas(data)
        rut = data.get("rut", "").strip()
        movimientos = self._build_result(fi, fc, rut)

        wb = Workbook()
        ws = wb.active
        ws.title = "OC/AT Pendientes"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["N° OC", "Fecha", "RUT", "Proveedor", "Artículo", "Cantidad", "P.Unit", "Total"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        sub_font = Font(name="Calibri", bold=True, size=10)
        sub_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        for m in movimientos:
            if m.get("_subtotal"):
                ws.append([
                    f"Subtotal {m['_numero']}", "", "", m["_proveedor"], "",
                    m["_cantidad"], "", m["_total"],
                ])
                row_num = ws.max_row
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = sub_font
                    cell.fill = sub_fill
                    cell.border = thin_border
                    cell.alignment = center_align
            else:
                art_label = f"{m['codigo']} {m['articulo']}".strip()
                ws.append([
                    m["numero"],
                    m["fecha"],
                    m["rut"],
                    m["proveedor"],
                    art_label,
                    m["cantidad"],
                    m["punit"],
                    m["total"],
                ])
                row_num = ws.max_row
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = center_align

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"ocat_pendientes_{rut or 'todos'}.xlsx"
        response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # --- PDF Resumen ---
    def _generar_pdf_resumen(self, request: HttpRequest, data: dict) -> HttpResponse:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        fi, fc = self._parse_fechas(data)
        rut = data.get("rut", "").strip()
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        resumen = self._build_resumen(fi, fc, rut)

        logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/brand-logos/logo-home-grande.png")
        if not os.path.exists(logo_path):
            logo_path = None

        prov_nombre = ""
        if rut:
            try:
                prov = Provclientes.objects.get(rut=rut)
                prov_nombre = prov.nombre
            except Provclientes.DoesNotExist:
                pass

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4*mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=2*mm, fontSize=8)
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=7, leading=10)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)

            elems = []
            elems.append(Paragraph("OC/AT Pendientes - Resumen", title_style))
            if rut:
                elems.append(Paragraph(f"<b>RUT:</b> {rut} &nbsp;&nbsp;&nbsp;<b>Nombre:</b> {prov_nombre}", subtitle_style))
            else:
                elems.append(Paragraph("<b>Todos los proveedores</b>", subtitle_style))

            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))
            elems.append(Spacer(1, 2*mm))

            headers = ["N° OC", "Fecha", "RUT", "Proveedor", "Cant.Total", "Cant.Suma", "Faltante"]
            col_widths = [22*mm, 20*mm, 22*mm, 50*mm, 22*mm, 24*mm, 26*mm]
            table_data = [headers]

            total_canttotal = 0
            total_cantsum = 0
            for r in resumen:
                faltante = r["faltante"]
                faltante_color = "red" if faltante != 0 else "green"
                table_data.append([
                    Paragraph(r["numero"], center_style),
                    Paragraph(r["fecha"], center_style),
                    Paragraph(r["rut"], center_style),
                    Paragraph(r["proveedor"], cell_style),
                    Paragraph(self._clq(r["canttotal"]), right_style),
                    Paragraph(self._clq(r["cant_sum"]), right_style),
                    Paragraph(f'<font color="{faltante_color}"><b>{self._clq(faltante)}</b></font>', right_style),
                ])
                total_canttotal += r["canttotal"]
                total_cantsum += r["cant_sum"]

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]))
            elems.append(tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w/2, 10*mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w-15*mm, 10*mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=18*mm, bottomMargin=16*mm)
        doc.build(build_elements(), canvasmaker=_Canvas)
        pdf_bytes = buf.getvalue()
        buf.close()
        filename = f"ocat_pendientes_resumen_{rut or 'todos'}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    # --- Excel Resumen ---
    def _generar_excel_resumen(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        fi, fc = self._parse_fechas(data)
        rut = data.get("rut", "").strip()
        resumen = self._build_resumen(fi, fc, rut)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["N° OC", "Fecha", "RUT", "Proveedor", "Cant.Total", "Cant.Suma", "Faltante"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        red_font = Font(name="Calibri", size=10, color="FF0000", bold=True)
        green_font = Font(name="Calibri", size=10, color="008000", bold=True)

        for r in resumen:
            ws.append([
                r["numero"], r["fecha"], r["rut"], r["proveedor"],
                r["canttotal"], r["cant_sum"], r["faltante"],
            ])
            row_num = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_align
            faltante_cell = ws.cell(row=row_num, column=7)
            faltante_cell.font = red_font if r["faltante"] != 0 else green_font

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"ocat_pendientes_resumen_{rut or 'todos'}.xlsx"
        response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
