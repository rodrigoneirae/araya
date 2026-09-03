import io
import os
from typing import Any
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.articulos import Articulos
from modulos.maestros.models.bodegas import Bodegas
from modulos.maestros.models.prov_cliente import Provclientes
from modulos.maestros.models.empleados import Empleados


class IndexInformeOrProvedorView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/inventario/informes/or_provedor.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        handlers = {
            "listar_proveedores": lambda _: self._listar_proveedores(),
            "buscar_proveedor": lambda d: self._buscar_proveedor(d.get("rut")),
            "info_articulo": lambda d: self._info_articulo(d),
            "info_ocat": lambda d: self._info_ocat(d),
            "informe_mensual": lambda d: self._informe_mensual(d),
            "generar_pdf_info": lambda d: self._generar_pdf_info(request, d),
            "generar_excel_info": lambda d: self._generar_excel_info(request, d),
            "generar_pdf_ocat": lambda d: self._generar_pdf_ocat(request, d),
            "generar_excel_ocat": lambda d: self._generar_excel_ocat(request, d),
            "generar_pdf_mensual": lambda d: self._generar_pdf_mensual(request, d),
            "generar_excel_mensual": lambda d: self._generar_excel_mensual(request, d),
        }
        handler = handlers.get(action, lambda _: JsonResponse({"success": False}))
        return handler(request.POST)

    def _listar_proveedores(self) -> JsonResponse:
        proveedores = (
            Provclientes.objects
            .filter(tipo__in=["Proveedor", "Ambos"])
            .values("rut", "nombre")
            .order_by("nombre")
        )
        return JsonResponse({"proveedores": list(proveedores)})

    def _buscar_proveedor(self, rut: str | None) -> JsonResponse:
        if not rut:
            return JsonResponse({"success": False})
        try:
            prov = Provclientes.objects.get(rut=rut.strip())
            return JsonResponse({
                "success": True,
                "data": {
                    "rut": prov.rut,
                    "nombre": prov.nombre,
                }
            })
        except Provclientes.DoesNotExist:
            return JsonResponse({"success": False, "message": "Proveedor no encontrado"})

    def _parse_fechas(self, data: dict) -> tuple:
        fi = None
        fc = None
        fecha_inicio = data.get("fecha_inicio", "").strip()
        fecha_corte = data.get("fecha_corte", "").strip()
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            except Exception:
                pass
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte, "%Y-%m-%d")
            except Exception:
                pass
        return fi, fc

    def _agrupar_por_codigo(self, qs, bodegas_map: dict) -> list[dict]:
        movimientos = []
        cod_actual = None
        sub_cant_entradas = 0
        sub_cant_salidas = 0
        sub_monto = 0

        def _flush_subtotal():
            nonlocal sub_cant_entradas, sub_cant_salidas, sub_monto
            saldo = sub_cant_entradas - sub_cant_salidas
            movimientos.append({
                "_subtotal": True,
                "_codigo": cod_actual,
                "_entradas": sub_cant_entradas,
                "_salidas": sub_cant_salidas,
                "_saldo": saldo,
                "_monto": sub_monto,
            })
            sub_cant_entradas = 0
            sub_cant_salidas = 0
            sub_monto = 0

        for m in qs.iterator():
            if not m.codigo:
                continue
            if m.codigo.codigo != cod_actual:
                if cod_actual is not None:
                    _flush_subtotal()
                cod_actual = m.codigo.codigo

            signo = m.tipo.signo if m.tipo else 0
            cantidad = m.cantidad or 0
            if signo > 0:
                sub_cant_entradas += cantidad
            if signo < 0:
                sub_cant_salidas += abs(cantidad)
            monto = cantidad * (m.punit or 0)
            sub_monto += monto

            movimientos.append({
                "_subtotal": False,
                "fecha": m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                "codigo": m.codigo.codigo,
                "nombre": m.codigo.descr or "",
                "um": m.codigo.um or "",
                "cantidad": cantidad,
                "punit": m.punit or 0,
                "total": monto,
                "tipo_nombre": m.tipo.nombre if m.tipo else "",
                "numero": str(int(m.numero)) if m.numero else "",
                "bodega_nombre": bodegas_map.get(m.bodega, str(m.bodega or "")),
                "signo": signo,
            })

        if cod_actual is not None:
            _flush_subtotal()

        return movimientos

    def _info_articulo(self, data: dict) -> JsonResponse:
        rut = data.get("rut", "").strip()
        if not rut:
            return JsonResponse({"success": False, "message": "RUT requerido"})
        fi, fc = self._parse_fechas(data)

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(rut=rut, tipo__isnull=False, codigo__isnull=False)
            .exclude(codigo__tipo='Servicio')
            .exclude(codigo__codigo='')
        )
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("codigo__codigo", "fecha")

        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}

        movimientos = self._agrupar_por_codigo(qs, bodegas_map)

        return JsonResponse({"success": True, "data": movimientos})

    def _info_ocat(self, data: dict) -> JsonResponse:
        rut = data.get("rut", "").strip()
        if not rut:
            return JsonResponse({"success": False, "message": "RUT requerido"})
        fi, fc = self._parse_fechas(data)

        encabezados = (
            Movs.objects
            .select_related("codigo")
            .filter(rut=rut, tipo=7, linea=0)
        )
        if fi:
            encabezados = encabezados.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            encabezados = encabezados.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        encabezados = encabezados.order_by("-numero")

        empleados_map = {}
        for e in Empleados.objects.values("cod", "nombre"):
            empleados_map[float(e["cod"])] = e["nombre"]

        numeros_ocat = [enc.numero for enc in encabezados]
        detalles_qs = Movs.objects.select_related("codigo").filter(
            rut=rut, tipo=7, numero__in=numeros_ocat, linea__gt=0
        ).order_by("numero", "linea")

        detalles_map = {}
        for d in detalles_qs:
            num = d.numero
            if num not in detalles_map:
                detalles_map[num] = []
            detalles_map[num].append({
                "linea": int(d.linea) if d.linea else 0,
                "codigo": d.codigo.codigo if d.codigo else "",
                "nombre": d.codigo.descr if d.codigo else "",
                "cantidad": float(d.cantidad or 0),
                "punit": float(d.punit or 0),
                "total": float(d.total or 0),
            })

        resultado = []
        for enc in encabezados.iterator():
            numero = enc.numero
            detalles = detalles_map.get(numero, [])
            encargado_nombre = empleados_map.get(float(enc.codencargado), "") if enc.codencargado else ""
            resultado.append({
                "numero": int(numero) if numero else "",
                "fecha": enc.fecha.strftime("%d-%m-%Y") if enc.fecha else "",
                "estado": enc.estado or "",
                "neto": float(enc.neto or 0),
                "total": float(enc.canttotal or 0) * float(enc.punit or 0),
                "encargado_nombre": encargado_nombre,
                "docref": str(int(enc.docref)) if enc.docref else "",
                "detalles": detalles,
            })

        return JsonResponse({"success": True, "data": resultado})

    def _informe_mensual(self, data: dict) -> JsonResponse:
        rut = data.get("rut", "").strip()
        ano_raw = data.get("ano", "").strip()
        if not rut or not ano_raw:
            return JsonResponse({"success": False, "message": "RUT y Año requeridos"})

        try:
            ano = int(ano_raw)
        except ValueError:
            return JsonResponse({"success": False, "message": "Año inválido"})

        detalles = (
            Movs.objects
            .select_related("codigo")
            .filter(rut=rut, tipo=7, linea__gt=0, fecha__year=ano)
            .order_by("codigo__codigo", "fecha")
        )

        articulos = {}
        for d in detalles.iterator():
            if not d.codigo:
                continue
            cod = d.codigo.codigo
            if cod not in articulos:
                articulos[cod] = {
                    "codigo": cod,
                    "nombre": d.codigo.descr or "",
                    "meses": {m: {"cant": 0, "valor": 0} for m in range(1, 13)},
                }
            mes = d.fecha.month
            cant = float(d.cantidad or 0)
            articulos[cod]["meses"][mes]["cant"] += cant
            articulos[cod]["meses"][mes]["valor"] += round(cant * float(d.punit or 0), 0)

        resultado = []
        for cod in sorted(articulos.keys()):
            art = articulos[cod]
            row = {"codigo": art["codigo"], "nombre": art["nombre"]}
            tot_cant = 0
            tot_valor = 0
            for m in range(1, 13):
                cant = round(art["meses"][m]["cant"], 3)
                valor = int(art["meses"][m]["valor"])
                row[f"m{m}_cant"] = cant
                row[f"m{m}_valor"] = valor
                tot_cant += cant
                tot_valor += valor
            row["tot_cant"] = round(tot_cant, 3)
            row["tot_valor"] = tot_valor
            resultado.append(row)

        return JsonResponse({"success": True, "data": resultado})

    def _generar_pdf_info(self, request: HttpRequest, data: dict) -> HttpResponse:
        rut = data.get("rut", "").strip()
        if not rut:
            return JsonResponse({"success": False, "message": "RUT requerido"})

        fi, fc = self._parse_fechas(data)
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        prov_nombre = ""
        try:
            prov = Provclientes.objects.get(rut=rut)
            prov_nombre = prov.nombre
        except Provclientes.DoesNotExist:
            pass

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(rut=rut, tipo__isnull=False, codigo__isnull=False)
            .exclude(codigo__tipo='Servicio')
            .exclude(codigo__codigo='')
        )
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("codigo__codigo", "fecha")

        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}

        grouped = self._agrupar_por_codigo(qs, bodegas_map)

        total_entradas = 0
        total_salidas = 0
        total_monto = 0
        for item in grouped:
            if item["_subtotal"]:
                total_entradas += item["_entradas"]
                total_salidas += item["_salidas"]
                total_monto += item["_monto"]

        saldo_final = total_entradas - total_salidas

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        subtotal_style_pdf = ParagraphStyle(
            "Subtotal", parent=getSampleStyleSheet()["Normal"],
            fontSize=7, leading=9, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#4b5563"),
        )
        subtotal_right = ParagraphStyle("SubR", parent=subtotal_style_pdf, alignment=2)

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=6, leading=9,
            )
            right_style = ParagraphStyle(
                "RightStyle", parent=cell_style, alignment=2,
            )
            center_style = ParagraphStyle(
                "CenterStyle", parent=cell_style, alignment=1,
            )
            resumen_style = ParagraphStyle(
                "Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1,
            )

            elems = []
            elems.append(Paragraph("Info Artículo por Proveedor", title_style))
            elems.append(Paragraph(
                f"<b>RUT:</b> {rut} &nbsp;&nbsp;&nbsp;"
                f"<b>Nombre:</b> {prov_nombre}",
                subtitle_style,
            ))
            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 3 * mm))

            header = ["Fecha", "Código", "Nombre", "UM", "Cantidad", "P.Unit", "Total", "Número", "Bodega"]
            table_data = [header]
            subtotal_rows = []

            for item in grouped:
                if item["_subtotal"]:
                    table_data.append([
                        Paragraph("", subtotal_style_pdf),
                        Paragraph(f"<b>Subtotal</b>", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph(f"<b>{clq(item['_saldo'])}</b>", subtotal_right),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph(f"<b>{cl(item['_monto'])}</b>", subtotal_right),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                    ])
                    subtotal_rows.append(len(table_data) - 1)
                else:
                    table_data.append([
                        Paragraph(str(item["fecha"]), center_style),
                        Paragraph(str(item["codigo"]), center_style),
                        Paragraph(str(item["nombre"]), cell_style),
                        Paragraph(str(item["um"]), center_style),
                        Paragraph(f"<b>{clq(item['cantidad'])}</b>", right_style),
                        Paragraph(cl(item["punit"]), right_style),
                        Paragraph(cl(item["total"]), right_style),
                        Paragraph(str(item["numero"]), center_style),
                        Paragraph(str(item["bodega_nombre"]), cell_style),
                    ])

            col_widths = [22*mm, 18*mm, 36*mm, 12*mm, 18*mm, 18*mm, 20*mm, 18*mm, 26*mm]

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]
            for sr in subtotal_rows:
                style_cmds.append(("BACKGROUND", (0, sr), (-1, sr), colors.HexColor("#e5e7eb")))
                style_cmds.append(("FONTNAME", (0, sr), (-1, sr), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)
            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[
                Paragraph(f'<b>Total Entradas</b><br/><font color="#16a34a">{clq(total_entradas)}</font>', resumen_style),
                Paragraph(f'<b>Total Salidas</b><br/><font color="#dc2626">{clq(total_salidas)}</font>', resumen_style),
                Paragraph(f'<b>Saldo</b><br/><font color="#2563eb">{clq(saldo_final)}</font>', resumen_style),
                Paragraph(f'<b>Monto Total</b><br/><font color="#7c3aed">{cl(total_monto)}</font>', resumen_style),
            ]]

            resumen_tbl = Table(resumen_data, colWidths=[32*mm, 32*mm, 32*mm, 32*mm], hAlign="CENTER")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elems.append(resumen_tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w/2, 10*mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w-15*mm, 10*mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=22*mm, bottomMargin=18*mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = f"info_articulo_{rut}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    def _generar_excel_info(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        rut = data.get("rut", "").strip()
        if not rut:
            return JsonResponse({"success": False, "message": "RUT requerido"})

        fi, fc = self._parse_fechas(data)

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(rut=rut, tipo__isnull=False, codigo__isnull=False)
            .exclude(codigo__tipo='Servicio')
            .exclude(codigo__codigo='')
        )
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("codigo__codigo", "fecha")
        bodegas_map = {b["cod"]: b["nombre"] for b in Bodegas.objects.values("cod", "nombre")}

        grouped = self._agrupar_por_codigo(qs, bodegas_map)

        wb = Workbook()
        ws = wb.active
        ws.title = "Info Artículo"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Fecha", "Código", "Nombre", "UM", "Cantidad", "P.Unit", "Total", "Tipo", "Número", "Bodega"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        sub_font = Font(name="Calibri", bold=True, size=10)
        sub_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        total_entradas = 0
        total_salidas = 0
        total_monto = 0

        for item in grouped:
            if item["_subtotal"]:
                total_entradas += item["_entradas"]
                total_salidas += item["_salidas"]
                total_monto += item["_monto"]

                sub_row = ["", f"Subtotal {item['_codigo']}", "", "", item["_saldo"], "", item["_monto"], "", "", ""]
                ws.append(sub_row)
                row_num = ws.max_row
                for col_idx in range(1, len(sub_row) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = sub_font
                    cell.fill = sub_fill
                    cell.border = thin_border
                    if col_idx in (5, 6, 7):
                        cell.alignment = num_align
                        cell.number_format = '#,##0.000' if col_idx == 5 else '#,##0'
                    else:
                        cell.alignment = center_align
            else:
                row_data = [
                    item["fecha"],
                    item["codigo"],
                    item["nombre"],
                    item["um"],
                    item["cantidad"],
                    item["punit"],
                    item["total"],
                    item["tipo_nombre"],
                    item["numero"],
                    item["bodega_nombre"],
                ]
                ws.append(row_data)
                row_num = ws.max_row
                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in (5, 6, 7):
                        cell.alignment = num_align
                        if col_idx == 5:
                            cell.number_format = '#,##0.000'
                        else:
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = center_align

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 22

        ws.append([])
        saldo_final = total_entradas - total_salidas

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        ws.append([f"Total Entradas: {clq(total_entradas)}  |  Total Salidas: {clq(total_salidas)}  |  Saldo: {clq(saldo_final)}  |  Monto Total: {cl(total_monto)}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"info_articulo_{rut}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf_ocat(self, request: HttpRequest, data: dict) -> HttpResponse:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        rut = data.get("rut", "").strip()
        if not rut:
            return JsonResponse({"success": False, "message": "RUT requerido"})

        fi, fc = self._parse_fechas(data)
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        prov_nombre = ""
        try:
            prov = Provclientes.objects.get(rut=rut)
            prov_nombre = prov.nombre
        except Provclientes.DoesNotExist:
            pass

        encabezados = (
            Movs.objects
            .select_related("codigo")
            .filter(rut=rut, tipo=7, linea=0)
        )
        if fi:
            encabezados = encabezados.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            encabezados = encabezados.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        encabezados = list(encabezados.order_by("fecha", "numero").values(
            "numero", "fecha", "estado", "neto", "canttotal", "punit", "codencargado", "docref"
        ))

        empleados_map = {}
        for e in Empleados.objects.values("cod", "nombre"):
            empleados_map[float(e["cod"])] = e["nombre"]

        numeros_ocat = [enc["numero"] for enc in encabezados]
        detalles_qs = Movs.objects.select_related("codigo").filter(
            rut=rut, tipo=7, numero__in=numeros_ocat, linea__gt=0
        ).order_by("numero", "linea").values(
            "numero", "linea", "codigo__codigo", "codigo__descr", "cantidad", "punit", "total", "bodega"
        )

        detalles_map = {}
        for d in detalles_qs:
            num = d["numero"]
            if num not in detalles_map:
                detalles_map[num] = []
            detalles_map[num].append({
                "linea": int(d["linea"]) if d["linea"] else 0,
                "codigo": d["codigo__codigo"] or "",
                "nombre": d["codigo__descr"] or "",
                "cantidad": float(d["cantidad"] or 0),
                "punit": float(d["punit"] or 0),
                "total": float(d["total"] or 0),
                "bodega": d["bodega"] or "",
            })

        resultado = []
        for enc in encabezados:
            numero = enc["numero"]
            detalles = detalles_map.get(numero, [])
            total_ocat = sum(d["cantidad"] for d in detalles)
            resultado.append({
                "numero": int(numero) if numero else "",
                "fecha": enc["fecha"].strftime("%d-%m-%Y") if enc["fecha"] else "",
                "neto": float(enc["neto"] or 0),
                "total": total_ocat,
                "docref": str(int(enc["docref"])) if enc["docref"] else "",
                "detalles": detalles,
            })

        total_neto = sum(o["neto"] for o in resultado)
        total_cantidad = sum(o["total"] for o in resultado)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4*mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=3*mm, fontSize=9,
            )
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=6, leading=8)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
            bold_style = ParagraphStyle("BoldStyle", parent=cell_style, fontName="Helvetica-Bold")

            elems = []
            elems.append(Paragraph("Informe OCAT por Proveedor (Con Detalle)", title_style))
            elems.append(Paragraph(
                f"<b>RUT:</b> {rut} &nbsp;&nbsp;&nbsp;"
                f"<b>Nombre:</b> {prov_nombre}",
                subtitle_style,
            ))
            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 3*mm))

            header_ocab = ["OCAT", "Fecha", "Neto", "Total Cant", "Doc.Ref"]
            header_det = ["L", "Código", "Nombre", "Cant", "P.Unit", "Total", "Bod"]

            for o in resultado:
                table_data_ocab = [header_ocab]
                table_data_ocab.append([
                    Paragraph(str(o["numero"]), bold_style),
                    Paragraph(o["fecha"], center_style),
                    Paragraph(cl(o["neto"]), right_style),
                    Paragraph(clq(o["total"]), right_style),
                    Paragraph(o["docref"], center_style),
                ])
                col_widths_ocab = [25*mm, 30*mm, 45*mm, 45*mm, 35*mm]
                tbl_ocab = Table(table_data_ocab, colWidths=col_widths_ocab)
                style_cmds_ocab = [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#e5e7eb")),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
                tbl_ocab.setStyle(TableStyle(style_cmds_ocab))
                elems.append(tbl_ocab)

                if o["detalles"]:
                    table_data_det = [header_det]
                    for det in o["detalles"]:
                        bodega_nom = ""
                        if det["bodega"]:
                            try:
                                bodega_nom = Bodegas.objects.get(cod=det["bodega"]).nombre
                            except Bodegas.DoesNotExist:
                                bodega_nom = str(det["bodega"])
                        table_data_det.append([
                            Paragraph(str(det["linea"]), center_style),
                            Paragraph(det["codigo"], center_style),
                            Paragraph(det["nombre"][:30] if det["nombre"] else "", cell_style),
                            Paragraph(clq(det["cantidad"]), right_style),
                            Paragraph(cl(det["punit"]), right_style),
                            Paragraph(cl(det["total"]), right_style),
                            Paragraph(bodega_nom, center_style),
                        ])
                    col_widths_det = [10*mm, 20*mm, 45*mm, 18*mm, 20*mm, 22*mm, 20*mm]
                    tbl_det = Table(table_data_det, colWidths=col_widths_det)
                    style_cmds_det = [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6b7280")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("FONTSIZE", (0, 0), (-1, -1), 5),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ]
                    tbl_det.setStyle(TableStyle(style_cmds_det))
                    elems.append(tbl_det)

                elems.append(Spacer(1, 4*mm))

            elems.append(Paragraph(
                f"<b>Total OCAT:</b> {len(resultado)} &nbsp;&nbsp;&nbsp;"
                f"<b>Neto Total:</b> ${cl(total_neto)} &nbsp;&nbsp;&nbsp;"
                f"<b>Total Cantidad:</b> {clq(total_cantidad)}",
                subtitle_style,
            ))
            elems.append(Spacer(1, 3*mm))
            elems.append(Paragraph(f"<i>Generado por {usuario} el {now_str}</i>", ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.gray)))

            return elems

        doc.build(build_elements())
        buffer.seek(0)

        filename = f"info_ocat_{rut}.pdf"
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_excel_ocat(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        rut = data.get("rut", "").strip()
        if not rut:
            return JsonResponse({"success": False, "message": "RUT requerido"})

        fi, fc = self._parse_fechas(data)

        encabezados = (
            Movs.objects
            .select_related("codigo")
            .filter(rut=rut, tipo=7, linea=0)
        )
        if fi:
            encabezados = encabezados.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            encabezados = encabezados.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        encabezados = list(encabezados.order_by("fecha", "numero").values(
            "numero", "fecha", "estado", "neto", "canttotal", "punit", "codencargado", "docref"
        ))

        empleados_map = {}
        for e in Empleados.objects.values("cod", "nombre"):
            empleados_map[float(e["cod"])] = e["nombre"]

        numeros_ocat = [enc["numero"] for enc in encabezados]
        detalles_qs = Movs.objects.select_related("codigo").filter(
            rut=rut, tipo=7, numero__in=numeros_ocat, linea__gt=0
        ).order_by("numero", "linea").values(
            "numero", "linea", "codigo__codigo", "codigo__descr", "cantidad", "punit", "total", "bodega"
        )

        detalles_map = {}
        for d in detalles_qs:
            num = d["numero"]
            if num not in detalles_map:
                detalles_map[num] = []
            detalles_map[num].append({
                "linea": int(d["linea"]) if d["linea"] else 0,
                "codigo": d["codigo__codigo"] or "",
                "nombre": d["codigo__descr"] or "",
                "cantidad": float(d["cantidad"] or 0),
                "punit": float(d["punit"] or 0),
                "total": float(d["total"] or 0),
                "bodega": d["bodega"] or "",
            })

        resultado = []
        for enc in encabezados:
            numero = enc["numero"]
            detalles = detalles_map.get(numero, [])
            encargado_nombre = empleados_map.get(float(enc["codencargado"]), "") if enc["codencargado"] else ""
            total_ocat = float(enc["canttotal"] or 0) * float(enc["punit"] or 0)
            resultado.append({
                "numero": int(numero) if numero else "",
                "fecha": enc["fecha"].strftime("%d-%m-%Y") if enc["fecha"] else "",
                "estado": enc["estado"] or "",
                "neto": float(enc["neto"] or 0),
                "total": total_ocat,
                "encargado_nombre": encargado_nombre,
                "docref": str(int(enc["docref"])) if enc["docref"] else "",
                "detalles": detalles,
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Info OCAT"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        hdr_det_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        hdr_det_fill = PatternFill(start_color="6b7280", end_color="6b7280", fill_type="solid")

        headers_ocab = ["OCAT", "Fecha", "Estado", "Neto", "Total", "Encargado", "Doc.Ref"]
        ws.append(headers_ocab)
        for col_idx in range(1, len(headers_ocab) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        sub_font = Font(name="Calibri", bold=True, size=9)
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        total_neto = 0
        total_monto = 0

        for o in resultado:
            total_neto += o["neto"]
            total_monto += o["total"]

            row_ocab = [
                o["numero"],
                o["fecha"],
                o["estado"],
                o["neto"],
                o["total"],
                o["encargado_nombre"],
                o["docref"],
            ]
            ws.append(row_ocab)
            row_num = ws.max_row
            for col_idx in range(1, len(row_ocab) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
                cell.border = thin_border
                if col_idx in (4, 5):
                    cell.alignment = num_align
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = center_align

            if o["detalles"]:
                headers_det = ["L", "Código", "Nombre", "Cantidad", "P.Unit", "Total", "Bodega"]
                ws.append(headers_det)
                row_num = ws.max_row
                for col_idx in range(1, len(headers_det) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = hdr_det_font
                    cell.fill = hdr_det_fill
                    cell.alignment = hdr_align
                    cell.border = thin_border

                for det in o["detalles"]:
                    bodega_nom = ""
                    if det["bodega"]:
                        try:
                            bodega_nom = Bodegas.objects.get(cod=det["bodega"]).nombre
                        except Bodegas.DoesNotExist:
                            bodega_nom = str(det["bodega"])

                    row_det = [
                        det["linea"],
                        det["codigo"],
                        det["nombre"],
                        det["cantidad"],
                        det["punit"],
                        det["total"],
                        bodega_nom,
                    ]
                    ws.append(row_det)
                    row_num = ws.max_row
                    for col_idx in range(1, len(row_det) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.font = data_font
                        cell.border = thin_border
                        if col_idx in (4, 5, 6):
                            cell.alignment = num_align
                            cell.number_format = '#,##0.000' if col_idx in (4,) else '#,##0'
                        else:
                            cell.alignment = center_align

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 28
        ws.column_dimensions["G"].width = 14

        if len(resultado) > 0 and resultado[0].get("detalles"):
            ws.column_dimensions["H"].width = 8
            ws.column_dimensions["I"].width = 14
            ws.column_dimensions["J"].width = 35
            ws.column_dimensions["K"].width = 14
            ws.column_dimensions["L"].width = 14
            ws.column_dimensions["M"].width = 16
            ws.column_dimensions["N"].width = 16

        ws.append([])
        ws.append([f"Total OCAT: {len(resultado)}  |  Neto Total: ${total_neto:,.0f}  |  Monto Total: ${total_monto:,.0f}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"info_ocat_{rut}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf_mensual(self, request: HttpRequest, data: dict) -> HttpResponse:
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        rut = data.get("rut", "").strip()
        ano_raw = data.get("ano", "").strip()
        if not rut or not ano_raw:
            return JsonResponse({"success": False, "message": "RUT y Año requeridos"})

        ano = int(ano_raw)

        prov_nombre = ""
        try:
            prov = Provclientes.objects.get(rut=rut)
            prov_nombre = prov.nombre
        except Provclientes.DoesNotExist:
            pass

        nombres_meses_abr = [
            "Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"
        ]

        # Query data (same as _informe_mensual)
        detalles = (
            Movs.objects
            .select_related("codigo")
            .filter(rut=rut, tipo=7, linea__gt=0, fecha__year=ano)
            .order_by("codigo__codigo", "fecha")
        )

        articulos = {}
        for d in detalles.iterator():
            if not d.codigo:
                continue
            cod = d.codigo.codigo
            if cod not in articulos:
                articulos[cod] = {
                    "codigo": cod,
                    "nombre": d.codigo.descr or "",
                    "meses": {m: {"cant": 0, "valor": 0} for m in range(1, 13)},
                }
            mes = d.fecha.month
            cant = float(d.cantidad or 0)
            articulos[cod]["meses"][mes]["cant"] += cant
            articulos[cod]["meses"][mes]["valor"] += round(cant * float(d.punit or 0), 0)

        data_rows = []
        for cod in sorted(articulos.keys()):
            art = articulos[cod]
            cant_row = [art["codigo"], art["nombre"]]
            val_row = [art["codigo"], art["nombre"]]
            tot_cant = 0
            tot_valor = 0
            for m in range(1, 13):
                cant = round(art["meses"][m]["cant"], 3)
                valor = int(art["meses"][m]["valor"])
                cant_row.append(cant)
                val_row.append(valor)
                tot_cant += cant
                tot_valor += valor
            cant_row.append(round(tot_cant, 3))
            val_row.append(tot_valor)
            data_rows.append((cant_row, val_row))

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        def build_table(title_text, is_valor):
            styles = getSampleStyleSheet()
            cell_style = ParagraphStyle("Cell", fontSize=7, leading=9)
            right_style = ParagraphStyle("Right", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("Center", parent=cell_style, alignment=1)
            bold_style = ParagraphStyle("Bold", parent=cell_style, fontName="Helvetica-Bold")
            bold_right = ParagraphStyle("BoldRight", parent=bold_style, alignment=2)
            header_style = ParagraphStyle("Hdr", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)

            # Header row: Codigo, Nombre, months...
            hdr = [Paragraph("<b>Código</b>", header_style), Paragraph("<b>Nombre</b>", header_style)]
            for mn in nombres_meses_abr:
                hdr.append(Paragraph(f"<b>{mn}</b>", header_style))
            hdr.append(Paragraph("<b>Total</b>", header_style))

            table_data = [hdr]
            for cr, vr in data_rows:
                row = [Paragraph(str(cr[0]), center_style), Paragraph(str(cr[1]), cell_style)]
                src = vr if is_valor else cr
                for i in range(2, len(src)):
                    if not is_valor:
                        row.append(Paragraph(clq(src[i]), right_style))
                    else:
                        row.append(Paragraph(f"${cl(src[i])}", right_style))
                table_data.append(row)

            # Total row
            tot_row = [Paragraph("<b>TOTAL</b>", bold_style), Paragraph("", cell_style)]
            monthly_sums = []
            for m in range(1, 13):
                s = sum((vr if is_valor else cr)[m + 1] for cr, vr in data_rows)
                if not is_valor:
                    s = round(s, 3)
                    tot_row.append(Paragraph(f"<b>{clq(s)}</b>", bold_right))
                else:
                    s = int(s)
                    tot_row.append(Paragraph(f"<b>${cl(s)}</b>", bold_right))
                monthly_sums.append(s)
            grand = sum(monthly_sums)
            if not is_valor:
                tot_row.append(Paragraph(f"<b>{clq(grand)}</b>", bold_right))
            else:
                tot_row.append(Paragraph(f"<b>${cl(grand)}</b>", bold_right))
            table_data.append(tot_row)

            col_widths = [14*mm, 28*mm] + [12*mm] * 12 + [14*mm]
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e5e7eb")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]))
            return tbl

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=3*mm, fontSize=13)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3*mm, fontSize=8)

            elems = []
            elems.append(Paragraph("Informe Mensual por Proveedor", title_style))
            elems.append(Paragraph(
                f"<b>RUT:</b> {rut} &nbsp;&nbsp;&nbsp;"
                f"<b>Nombre:</b> {prov_nombre} &nbsp;&nbsp;&nbsp;"
                f"<b>Año:</b> {ano}",
                subtitle_style,
            ))

            # Quantity table
            tit_style = ParagraphStyle("TableTitle", parent=styles["Normal"], fontSize=9, leading=12, spaceAfter=2*mm, fontName="Helvetica-Bold")
            elems.append(Paragraph("Cantidades", tit_style))
            elems.append(build_table("Cantidades", is_valor=False))
            elems.append(Spacer(1, 5*mm))

            # Value table
            elems.append(Paragraph("Valores", tit_style))
            elems.append(build_table("Valores", is_valor=True))

            return elems

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=12*mm, rightMargin=12*mm,
            topMargin=18*mm, bottomMargin=16*mm,
        )
        doc.build(build_elements())

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = f"informe_mensual_{rut}_{ano}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    def _generar_excel_mensual(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        rut = data.get("rut", "").strip()
        ano_raw = data.get("ano", "").strip()
        if not rut or not ano_raw:
            return JsonResponse({"success": False, "message": "RUT y Año requeridos"})

        ano = int(ano_raw)

        nombres_meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]

        ocat_encabezados = (
            Movs.objects
            .filter(rut=rut, tipo=7, linea=0, fecha__year=ano)
            .order_by("fecha")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Mensual"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Mes", "OCAT", "Neto", "Monto"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        total_neto = 0
        total_monto = 0
        total_ocat = 0

        for mes_num in range(1, 13):
            ocat_mes = ocat_encabezados.filter(fecha__month=mes_num)
            neto_sum = 0
            total_sum = 0
            for enc in ocat_mes.iterator():
                neto_sum += float(enc.neto or 0)
                total_sum += float(enc.canttotal or 0) * float(enc.punit or 0)
            cnt = ocat_mes.count()
            total_neto += neto_sum
            total_monto += total_sum
            total_ocat += cnt

            row_data = [nombres_meses[mes_num-1], cnt, neto_sum, total_sum]
            ws.append(row_data)
            row_num = ws.max_row
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in (2, 3, 4):
                    cell.alignment = num_align
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = center_align

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        ws.append([])

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")

        ws.append([f"Total OCAT: {total_ocat}  |  Neto Anual: {cl(total_neto)}  |  Monto Anual: {cl(total_monto)}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"informe_mensual_{rut}_{ano}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context
