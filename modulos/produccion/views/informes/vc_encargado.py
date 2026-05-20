import io
import os
from datetime import datetime
from typing import Any

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.empleados import Empleados
from modulos.maestros.models.articulos import Articulos


class IndexInformeVCEncargadoView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/produccion/informes/vc_encargado.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        if action == "listar_encargados":
            return self._listar_encargados()
        elif action == "listar_datos":
            return self._listar_datos(request)
        elif action == "generar_pdf":
            return self._generar_pdf(request)
        elif action == "generar_excel":
            return self._generar_excel(request)
        return JsonResponse({"success": False})

    def _listar_encargados(self) -> JsonResponse:
        empleados = Empleados.objects.values("cod", "nombre").order_by("nombre")
        return JsonResponse({"encargados": list(empleados)})

    def _listar_datos(self, request: HttpRequest) -> JsonResponse:
        fecha_inicio = request.POST.get("fecha_inicio", "").strip()
        fecha_fin = request.POST.get("fecha_fin", "").strip()
        encargado_cod = request.POST.get("encargado", "").strip()

        data = self._get_datos(fecha_inicio, fecha_fin, encargado_cod if encargado_cod else None)

        enc_data = None
        if data and isinstance(data, dict):
            enc_data = list(data.values())[0] if data else None

        rows = []
        if enc_data and isinstance(enc_data, dict):
            for art_cod, art_data in sorted(enc_data["articulos"].items()):
                for mov in art_data["movimientos"]:
                    rows.append({
                        "articulo": art_cod,
                        "nombre": art_data["nombre"],
                        "fecha": mov["fecha"],
                        "ot": mov["ot"] if mov["ot"] else "",
                        "tipo": mov["tipo"] or "",
                        "cantidad": mov["cantidad"],
                        "um": mov["um"],
                    })

        return JsonResponse({"data": rows})

    def _get_datos(self, fecha_inicio: str, fecha_fin: str, encargado_cod: str | None):
        fi = None
        ff = None
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d")
            except Exception:
                pass
        if fecha_fin:
            try:
                ff = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d")
            except Exception:
                pass

        qs = Movs.objects.select_related("codigo", "tipo").filter(tipo__cod__in=[10]).exclude(cantidad=0)

        if fi:
            qs = qs.filter(fecha__gte=fi)
        if ff:
            qs = qs.filter(fecha__lte=ff.replace(hour=23, minute=59, second=59))
        if encargado_cod:
            qs = qs.filter(codencargado=encargado_cod)

        qs = qs.order_by("codencargado", "codigo", "fecha", "numero")

        empleados_map = {e["cod"]: e["nombre"] for e in Empleados.objects.values("cod", "nombre")}

        data = {}
        for m in qs.iterator():
            enc_cod = m.codencargado or ""
            enc_nom = empleados_map.get(enc_cod, enc_cod)
            if enc_cod not in data:
                data[enc_cod] = {"nombre": enc_nom, "articulos": {}, "total_cantidad": 0}

            cod_art = m.codigo.codigo if m.codigo else ""
            nom_art = m.codigo.descr if m.codigo else ""
            um_art = m.codigo.um if m.codigo else ""
            ot_num = int(m.docref) if m.docref else 0
            fecha = m.fecha.strftime("%d-%m-%Y") if m.fecha else ""
            tipo_nom = m.codigo.tipo if m.codigo and m.codigo.tipo else ""

            if cod_art not in data[enc_cod]["articulos"]:
                data[enc_cod]["articulos"][cod_art] = {"nombre": nom_art, "um": um_art, "movimientos": [], "total_cantidad": 0}

            cantidad = m.cantidad if m.cantidad else 0
            if cantidad == 0:
                continue

            data[enc_cod]["articulos"][cod_art]["movimientos"].append({
                "fecha": fecha,
                "ot": ot_num,
                "tipo": tipo_nom,
                "cantidad": cantidad,
                "um": um_art,
            })
            data[enc_cod]["articulos"][cod_art]["total_cantidad"] += cantidad
            data[enc_cod]["total_cantidad"] += cantidad

        return data

    def _generar_excel(self, request: HttpRequest) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        fecha_inicio = request.POST.get("fecha_inicio", "").strip()
        fecha_fin = request.POST.get("fecha_fin", "").strip()
        encargado_cod = request.POST.get("encargado", "").strip()

        data = self._get_datos(fecha_inicio, fecha_fin, encargado_cod if encargado_cod else None)

        wb = Workbook()
        ws = wb.active
        ws.title = "Consumo por Encargado"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
)

        headers = ["Artículo", "Nombre", "Fecha", "N° OT", "Tipo", "Cant", "UM"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        enc_data = None
        if data and isinstance(data, dict):
            enc_data = list(data.values())[0] if data else None

        if enc_data and isinstance(enc_data, dict):
            for art_cod, art_data in sorted(enc_data["articulos"].items()):
                for mov in art_data["movimientos"]:
                    ws.append([art_cod, art_data["nombre"], mov["fecha"], mov["ot"], mov["tipo"], mov["cantidad"], mov["um"]])
                    row_num = ws.max_row
                    for col_idx in range(1, 8):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.border = thin_border
                        if col_idx == 6:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0.000'
                        elif col_idx in [1, 2]:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 8

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "consumo_encargado.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generar_pdf(self, request: HttpRequest) -> HttpResponse:
        fecha_inicio = request.POST.get("fecha_inicio", "").strip()
        fecha_fin = request.POST.get("fecha_fin", "").strip()
        encargado_cod = request.POST.get("encargado", "").strip()
        usuario = str(request.user)

        data = self._get_datos(fecha_inicio, fecha_fin, encargado_cod if encargado_cod else None)

        fi = None
        ff = None
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d")
            except Exception:
                pass
        if fecha_fin:
            try:
                ff = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d")
            except Exception:
                pass

        fecha_inf = ""
        if fi or ff:
            fecha_inf = "del "
            if fi:
                fecha_inf += fi.strftime("%d-%m-%Y")
            fecha_inf += " al "
            if ff:
                fecha_inf += ff.strftime("%d-%m-%Y")

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/brand-logos/logo-home-grande.png")
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")
        
        def clt(v):
            return f"{int(round(v)):,}".replace(",", ".")

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14)
            subtitle_style = ParagraphStyle("CustomSub", parent=styles["Normal"], spaceAfter=3 * mm, fontSize=9)
            cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=7, leading=9)
            right_style = ParagraphStyle("RightStyle", parent=cell_style, alignment=2)
            center_style = ParagraphStyle("CenterStyle", parent=cell_style, alignment=1)
            resumen_style = ParagraphStyle("Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1)

            elems = []
            elems.append(Paragraph("Informe de Consumos por Encargado", title_style))
            elems.append(Paragraph(f"<b>Encargado:</b> {list(data.values())[0]['nombre'] if data else ''}", subtitle_style))
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 3 * mm))

            header = ["Artículo", "Nombre", "Fecha", "N° OT", "Tipo", "Cant", "UM"]
            col_widths = [22 * mm, 50 * mm, 20 * mm, 16 * mm, 35 * mm, 18 * mm, 14 * mm]

            enc_data = list(data.values())[0] if data else None

            if not enc_data:
                return HttpResponse("No hay datos", content_type="text/plain")

            table_data = [header]

            for art_cod, art_data in sorted(enc_data["articulos"].items()):
                if not art_data["movimientos"]:
                    continue
                subtotal = 0
                for mov in art_data["movimientos"]:
                    table_data.append([
                        Paragraph(str(art_cod), center_style),
                        Paragraph(art_data["nombre"], cell_style),
                        Paragraph(mov["fecha"], center_style),
                        Paragraph(str(mov["ot"]) if mov["ot"] else "", center_style),
                        Paragraph(mov["tipo"] or "", cell_style),
                        Paragraph(clq(mov["cantidad"]), right_style),
                        Paragraph(mov["um"], center_style),
                    ])
                    subtotal += mov["cantidad"]
                
                table_data.append([
                    Paragraph("", center_style),
                    Paragraph("<b>Subtotal</b>", ParagraphStyle("SubTotal", parent=cell_style, fontSize=7, bold=True)),
                    Paragraph("", center_style),
                    Paragraph("", center_style),
                    Paragraph("", cell_style),
                    Paragraph(clt(subtotal), ParagraphStyle("SubTotalNum", parent=right_style, fontSize=7, bold=True)),
                    Paragraph("", center_style),
                ])

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("ALIGN", (2, 0), (4, -1), "CENTER"),
            ]
            
            for i in range(1, len(table_data)):
                if len(table_data[i]) > 1 and "Subtotal" in str(table_data[i][1]):
                    style_cmds.extend([
                        ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#e5e7eb")),
                        ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"),
                    ])
                else:
                    bg = colors.white if i % 2 == 1 else colors.HexColor("#f9fafb")
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
            
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []

            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15 * mm, h - 17 * mm, width=50 * mm, height=11 * mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w / 2, 10 * mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w - 15 * mm, 10 * mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=22 * mm, bottomMargin=18 * mm)
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = "consumo_encargado.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response