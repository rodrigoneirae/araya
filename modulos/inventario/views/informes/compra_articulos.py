import io
import os
from typing import Any
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.generic import TemplateView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from modulos.inventario.models.movs import Movs
from modulos.maestros.models.articulos import Articulos
from modulos.maestros.models.prov_cliente import Provclientes


class IndexInformeCompraArticulosView(LoginRequiredMixin, TemplateView):
    template_name = 'modulos/inventario/informes/compra_articulos.html'

    def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        action = request.POST.get("action", "")
        handlers = {
            "listar_articulos": lambda _: self._listar_articulos(),
            "buscar_articulo": lambda d: self._buscar_articulo(d.get("codigo")),
            "info_proveedor": lambda d: self._info_proveedor(d),
            "generar_pdf": lambda d: self._generar_pdf(request, d),
            "generar_excel": lambda d: self._generar_excel(request, d),
        }
        handler = handlers.get(action, lambda _: JsonResponse({"success": False}))
        return handler(request.POST)

    def _listar_articulos(self) -> JsonResponse:
        articulos = (
            Articulos.objects
            .exclude(codigo='')
            .values("codigo", "descr", "um")
            .order_by("descr")
        )
        return JsonResponse({"articulos": list(articulos)})

    def _buscar_articulo(self, codigo: str | None) -> JsonResponse:
        if not codigo:
            return JsonResponse({"success": False})
        try:
            art = Articulos.objects.get(codigo=codigo.strip())
            return JsonResponse({
                "success": True,
                "data": {
                    "codigo": art.codigo,
                    "descr": art.descr,
                    "um": art.um,
                }
            })
        except Articulos.DoesNotExist:
            return JsonResponse({"success": False, "message": "Artículo no encontrado"})

    def _parse_fechas(self, data: dict) -> tuple:
        fi = None
        fc = None
        fecha_inicio = data.get("fecha_inicio", "").strip()
        fecha_corte = data.get("fecha_corte", "").strip()
        if fecha_inicio:
            try:
                fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            except Exception:
                pass
        if fecha_corte:
            try:
                fc = datetime.strptime(fecha_corte, "%Y-%m-%d")
            except Exception:
                pass
        return fi, fc

    def _agrupar_por_articulo(self, qs, proveedores_map: dict) -> list[dict]:
        movimientos = []
        codigo_actual = None
        sub_cantidad = 0
        sub_monto = 0

        def _flush_subtotal():
            nonlocal sub_cantidad, sub_monto
            movimientos.append({
                "_subtotal": True,
                "_codigo": codigo_actual,
                "_cantidad": sub_cantidad,
                "_monto": sub_monto,
            })
            sub_cantidad = 0
            sub_monto = 0

        for m in qs.iterator():
            art_cod = m.codigo.codigo if m.codigo else ""
            if art_cod != codigo_actual:
                if codigo_actual is not None:
                    _flush_subtotal()
                codigo_actual = art_cod

            cantidad = m.cantidad or 0
            monto = cantidad * (m.punit or 0)
            sub_cantidad += cantidad
            sub_monto += monto

            prov_nombre = proveedores_map.get(m.rut, "")

            movimientos.append({
                "_subtotal": False,
                "articulo_codigo": art_cod,
                "articulo_nombre": m.codigo.descr if m.codigo else "",
                "fecha": m.fecha.strftime("%d-%m-%Y") if m.fecha else "",
                "rut": m.rut,
                "proveedor_nombre": prov_nombre,
                "cantidad": cantidad,
                "punit": m.punit or 0,
                "total": monto,
                "tipo_nombre": m.tipo.nombre if m.tipo else "",
                "numero": str(int(m.numero)) if m.numero else "",
            })

        if codigo_actual is not None:
            _flush_subtotal()

        return movimientos

    def _info_proveedor(self, data: dict) -> JsonResponse:
        codigo = data.get("codigo", "").strip()
        fi, fc = self._parse_fechas(data)

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(tipo=7, linea__gt=0)
        )
        if codigo:
            qs = qs.filter(codigo=codigo)
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("codigo", "-cantidad")

        proveedores_map = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_map[p["rut"]] = p["nombre"]

        movimientos = self._agrupar_por_articulo(qs, proveedores_map)

        total_cant = sum(m["_cantidad"] for m in movimientos if m.get("_subtotal"))
        total_monto = sum(m["_monto"] for m in movimientos if m.get("_subtotal"))
        for m in movimientos:
            if m.get("_subtotal"):
                m["_participacion_cant"] = round(m["_cantidad"] / total_cant * 100, 2) if total_cant else 0
                m["_participacion_monto"] = round(m["_monto"] / total_monto * 100, 2) if total_monto else 0

        return JsonResponse({"success": True, "data": movimientos})

    def _generar_pdf(self, request: HttpRequest, data: dict) -> HttpResponse:
        codigo = data.get("codigo", "").strip()
        fi, fc = self._parse_fechas(data)
        usuario = str(request.user)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        art_nombre = ""
        if codigo:
            try:
                art = Articulos.objects.get(codigo=codigo)
                art_nombre = art.descr
            except Articulos.DoesNotExist:
                pass

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(tipo=7, linea__gt=0)
        )
        if codigo:
            qs = qs.filter(codigo=codigo)
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("codigo", "-cantidad")

        proveedores_map = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_map[p["rut"]] = p["nombre"]

        movimientos = self._agrupar_por_articulo(qs, proveedores_map)

        total_cantidad = sum(m["_cantidad"] for m in movimientos if m.get("_subtotal"))
        total_monto = sum(m["_monto"] for m in movimientos if m.get("_subtotal"))
        for m in movimientos:
            if m.get("_subtotal"):
                m["_participacion_cant"] = round(m["_cantidad"] / total_cantidad * 100, 2) if total_cantidad else 0
                m["_participacion_monto"] = round(m["_monto"] / total_monto * 100, 2) if total_monto else 0

        logo_path = os.path.join(
            settings.STATIC_ROOT,
            "assets/images/brand-logos/logo-home-grande.png"
        )
        if not os.path.exists(logo_path):
            logo_path = None

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        subtotal_style_pdf = ParagraphStyle(
            "Subtotal", parent=getSampleStyleSheet()["Normal"],
            fontSize=6, leading=9, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#4b5563"),
        )
        subtotal_right = ParagraphStyle("SubR", parent=subtotal_style_pdf, alignment=2)

        def build_elements():
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading2"], spaceAfter=4 * mm, fontSize=14
            )
            subtitle_style = ParagraphStyle(
                "CustomSub", parent=styles["Normal"], spaceAfter=2 * mm, fontSize=8,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=5.5, leading=8,
            )
            right_style = ParagraphStyle(
                "RightStyle", parent=cell_style, alignment=2,
            )
            center_style = ParagraphStyle(
                "CenterStyle", parent=cell_style, alignment=1,
            )
            resumen_style = ParagraphStyle(
                "Resumen", parent=styles["Normal"], fontSize=8, leading=11, alignment=1,
            )

            elems = []
            elems.append(Paragraph("Compra por Artículo", title_style))
            if codigo:
                elems.append(Paragraph(
                    f"<b>Código:</b> {codigo} &nbsp;&nbsp;&nbsp;"
                    f"<b>Nombre:</b> {art_nombre}",
                    subtitle_style,
                ))
            else:
                elems.append(Paragraph("<b>Todos los artículos</b>", subtitle_style))
            fecha_inf = ""
            if fi and fc:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')} al {fc.strftime('%d-%m-%Y')}"
            elif fc:
                fecha_inf = f"Al {fc.strftime('%d-%m-%Y')}"
            elif fi:
                fecha_inf = f"Desde {fi.strftime('%d-%m-%Y')}"
            if fecha_inf:
                elems.append(Paragraph(f"<b>Período:</b> {fecha_inf}", subtitle_style))

            elems.append(Spacer(1, 2 * mm))

            header = ["Artículo", "Fecha", "RUT", "Proveedor", "Cant.", "Part.", "Total", "Número"]
            table_data = [header]
            subtotal_rows = []

            for item in movimientos:
                if item.get("_subtotal"):
                    table_data.append([
                        Paragraph(f"<b>Subtotal {item['_codigo']}</b>", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph("", subtotal_style_pdf),
                        Paragraph(f"<b>{clq(item['_cantidad'])}</b>", subtotal_right),
                        Paragraph(f"<b>{item['_participacion_cant']:.2f}%</b>", subtotal_right),
                        Paragraph(f"<b>{cl(item['_monto'])}</b>", subtotal_right),
                        Paragraph("", subtotal_style_pdf),
                    ])
                    subtotal_rows.append(len(table_data) - 1)
                else:
                    art_label = f"{item.get('articulo_codigo', '')} {item.get('articulo_nombre', '')}".strip()
                    table_data.append([
                        Paragraph(art_label, cell_style),
                        Paragraph(str(item["fecha"]), center_style),
                        Paragraph(str(item["rut"]), center_style),
                        Paragraph(str(item["proveedor_nombre"]), cell_style),
                        Paragraph(f"<b>{clq(item['cantidad'])}</b>", right_style),
                        Paragraph("", cell_style),
                        Paragraph(cl(item["total"]), right_style),
                        Paragraph(str(item["numero"]), center_style),
                    ])

            col_widths = [32*mm, 16*mm, 16*mm, 40*mm, 18*mm, 14*mm, 22*mm, 18*mm]

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]
            for sr in subtotal_rows:
                style_cmds.append(("BACKGROUND", (0, sr), (-1, sr), colors.HexColor("#e5e7eb")))
                style_cmds.append(("FONTNAME", (0, sr), (-1, sr), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(style_cmds))
            elems.append(tbl)
            elems.append(Spacer(1, 5 * mm))

            resumen_data = [[
                Paragraph(f'<b>Total Cantidad</b><br/><font color="#2563eb">{clq(total_cantidad)}</font>', resumen_style),
                Paragraph(f'<b>Monto Total</b><br/><font color="#7c3aed">{cl(total_monto)}</font>', resumen_style),
                Paragraph(f'<b>Total Registros</b><br/><font color="#4b5563">{len(movimientos)}</font>', resumen_style),
            ]]

            resumen_tbl = Table(resumen_data, colWidths=[42*mm, 42*mm, 42*mm], hAlign="CENTER")
            resumen_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d1d5db")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            elems.append(resumen_tbl)
            return elems

        class _Canvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved)
                for state in self._saved:
                    self.__dict__.update(state)
                    w, h = self._pagesize
                    if logo_path:
                        try:
                            from PIL import Image as PILImage
                            from reportlab.lib.utils import ImageReader
                            _img = PILImage.open(logo_path)
                            if _img.mode == "RGBA":
                                _bg = PILImage.new("RGB", _img.size, (255, 255, 255))
                                _bg.paste(_img, mask=_img.split()[3])
                                self.drawImage(ImageReader(_bg), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                            else:
                                self.drawImage(ImageReader(_img), 15*mm, h-17*mm, width=50*mm, height=11*mm, preserveAspectRatio=True)
                        except Exception:
                            pass
                    self.setFont("Helvetica", 7)
                    self.setFillColor(colors.HexColor("#6b7280"))
                    self.drawCentredString(w/2, 10*mm, f"Página {self.getPageNumber()} de {total}")
                    self.drawRightString(w-15*mm, 10*mm, f"Usuario: {usuario} - {now_str}")
                    super().showPage()
                super().save()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=12*mm, rightMargin=12*mm,
            topMargin=18*mm, bottomMargin=16*mm,
        )
        doc.build(build_elements(), canvasmaker=_Canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        filename = f"compra_articulos_{codigo or 'todos'}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    def _generar_excel(self, request: HttpRequest, data: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        codigo = data.get("codigo", "").strip()
        fi, fc = self._parse_fechas(data)

        qs = (
            Movs.objects
            .select_related("tipo", "codigo")
            .filter(tipo=7, linea__gt=0)
        )
        if codigo:
            qs = qs.filter(codigo=codigo)
        if fi:
            qs = qs.filter(fecha__gte=fi.replace(hour=0, minute=0, second=0))
        if fc:
            qs = qs.filter(fecha__lte=fc.replace(hour=23, minute=59, second=59))

        qs = qs.order_by("codigo", "-cantidad")

        proveedores_map = {}
        for p in Provclientes.objects.values("rut", "nombre"):
            proveedores_map[p["rut"]] = p["nombre"]

        movimientos = self._agrupar_por_articulo(qs, proveedores_map)

        total_cantidad = sum(m["_cantidad"] for m in movimientos if m.get("_subtotal"))
        total_monto = sum(m["_monto"] for m in movimientos if m.get("_subtotal"))
        for m in movimientos:
            if m.get("_subtotal"):
                m["_participacion_cant"] = round(m["_cantidad"] / total_cantidad * 100, 2) if total_cantidad else 0
                m["_participacion_monto"] = round(m["_monto"] / total_monto * 100, 2) if total_monto else 0

        wb = Workbook()
        ws = wb.active
        ws.title = "Compra por Artículo"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        headers = ["Artículo", "Fecha", "RUT", "Proveedor", "Cantidad", "Part.", "Total", "Tipo", "Número"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        data_font = Font(name="Calibri", size=10)
        sub_font = Font(name="Calibri", bold=True, size=10)
        sub_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
        num_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")

        total_cantidad = 0
        total_monto = 0

        for item in movimientos:
            if item.get("_subtotal"):
                total_cantidad += item["_cantidad"]
                total_monto += item["_monto"]

                sub_row = [f"Subtotal {item['_codigo']}", "", "", "", item["_cantidad"], f"{item['_participacion_cant']:.2f}%", item["_monto"], "", ""]
                ws.append(sub_row)
                row_num = ws.max_row
                for col_idx in range(1, len(sub_row) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = sub_font
                    cell.fill = sub_fill
                    cell.border = thin_border
                    if col_idx in (5, 7):
                        cell.alignment = num_align
                        cell.number_format = '#,##0.000' if col_idx == 5 else '#,##0'
                    else:
                        cell.alignment = center_align
            else:
                art_label = f"{item.get('articulo_codigo', '')} {item.get('articulo_nombre', '')}".strip()
                row_data = [
                    art_label,
                    item["fecha"],
                    item["rut"],
                    item["proveedor_nombre"],
                    item["cantidad"],
                    "",
                    item["total"],
                    item["tipo_nombre"],
                    item["numero"],
                ]
                ws.append(row_data)
                row_num = ws.max_row
                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in (5, 7):
                        cell.alignment = num_align
                        cell.number_format = '#,##0.000' if col_idx == 5 else '#,##0'
                    else:
                        cell.alignment = center_align

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 14

        ws.append([])

        def cl(v):
            return f"{int(round(float(v))):,}".replace(",", ".")
        def clq(v):
            val = float(v)
            return f"{int(round(val)):,}".replace(",", ".") if val == int(val) else f"{val:,.3f}".replace(",", ".")

        ws.append([f"Total Cantidad: {clq(total_cantidad)}  |  Monto Total: {cl(total_monto)}"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"compra_articulos_{codigo or 'todos'}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
